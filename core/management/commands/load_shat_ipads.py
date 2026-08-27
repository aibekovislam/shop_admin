from collections import OrderedDict
from decimal import Decimal, ROUND_HALF_UP
from io import BytesIO
import hashlib
import json
import re
from pathlib import Path
from urllib.parse import urlparse

from django.core.files.base import ContentFile
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook
from PIL import Image
import requests

from core.marketplace.factory import get_marketplace_adapter
from core.models import (
    Brand,
    BrandCategory,
    Channel,
    ChannelPrice,
    Memory,
    Product,
    ProductCategory,
    ProductColor,
    ProductImage,
    ProductSize,
    ProductVariant,
    Stock,
)


STORE_IMAGE_BASE = "https://store.storeimages.cdn-apple.com/1/as-images.apple.com/is/"


COLORS = {
    "BLUE": ("Blue", "#A7C9E8"),
    "PINK": ("Pink", "#F6C8D0"),
    "SILVER": ("Silver", "#E3E4E5"),
    "YELLOW": ("Yellow", "#F5D35C"),
    "SPACE GRAY": ("Space Gray", "#5D5E60"),
    "SPACEGRAY": ("Space Gray", "#5D5E60"),
    "SPACE BLACK": ("Space Black", "#2E2E30"),
    "S/BLACK": ("Space Black", "#2E2E30"),
    "BLACK": ("Black", "#1D1D1F"),
    "STARLIGHT": ("Starlight", "#F0E4D3"),
}


def store_image(image_id, width=1200, height=900, image_format="jpeg"):
    return f"{STORE_IMAGE_BASE}{image_id}?wid={width}&hei={height}&fmt={image_format}"


IPAD_A16_GALLERY = [
    store_image("ipad-model-unselect-gallery-1-202503"),
    store_image("ipad-model-unselect-gallery-2-202503"),
    store_image("ipad-model-unselect-gallery-3-202503"),
]


IPAD_AIR_GALLERY = [
    store_image("ipad-air-model-unselect-gallery-1-202405"),
    store_image("ipad-air-model-unselect-gallery-2-202405"),
    store_image("ipad-air-model-unselect-gallery-3-202405"),
]


IPAD_PRO_GALLERY = [
    store_image("ipad-pro-model-select-gallery-1-202405"),
    store_image("ipad-pro-finish-unselect-gallery-2-202405"),
    store_image("ipad-pro-model-select-gallery-2-202405"),
]


KEYBOARD_GALLERY = {
    "folio": [
        store_image("ipad-accessories-keyboard-card1-202210"),
        store_image("ipad-accessories-keyboard-card2-202210"),
        store_image("ipad-accessories-keyboard-card3-202210"),
    ],
    "air": [
        store_image("ipadair-accessory-magickeyboard-card1_GEO_US"),
        store_image("ipadair-accessory-magickeyboard-card2"),
        store_image("ipadair-accessory-magickeyboard-card3"),
    ],
    "pro": [
        store_image("ipad-pro-accessory-magic-keyboard-card1-202210_GEO_US"),
        store_image("ipad-pro-magic-keyboard-select-gallery-2-202405"),
        store_image("ipad-pro-accessory-magic-keyboard-card3-202210_GEO_US"),
    ],
}


class Command(BaseCommand):
    help = "Загружает iPad из Excel Turan в канал SHAT / M-Market."

    def add_arguments(self, parser):
        parser.add_argument(
            "--excel",
            default="Price Turan .xlsx",
            help="Путь к Excel-файлу с остатками Turan.",
        )
        parser.add_argument("--channel-id", type=int, help="ID канала M-Market.")
        parser.add_argument("--adapter-key", default="mmarket", help="Ключ адаптера канала.")
        parser.add_argument("--markup", type=Decimal, default=Decimal("0.15"), help="Наценка, например 0.15.")
        parser.add_argument("--usd-rate", type=Decimal, default=Decimal("88"), help="Курс USD/KGS.")
        parser.add_argument("--dry-run", action="store_true", help="Создать товары и показать payload без отправки.")
        parser.add_argument("--skip-images", action="store_true", help="Не скачивать фото, если их уже минимум 3.")

    def handle(self, *args, **options):
        channel = self.get_channel(options)
        rows = self.read_ipads(options["excel"])
        payload_ids = []

        with transaction.atomic():
            brand, _ = Brand.objects.get_or_create(name="Apple")
            tablet_category, _ = ProductCategory.objects.get_or_create(name="Планшеты")
            accessory_category, _ = ProductCategory.objects.get_or_create(name="Аксессуары для планшетов")
            ipad_brand_category, _ = BrandCategory.objects.get_or_create(brand=brand, name="iPad")
            accessory_brand_category, _ = BrandCategory.objects.get_or_create(brand=brand, name="iPad Accessories")

            for key, stock_row in rows.items():
                item = self.build_item(key)
                category = accessory_category if item["is_accessory"] else tablet_category
                brand_category = accessory_brand_category if item["is_accessory"] else ipad_brand_category
                price = self.calculate_price(stock_row["usd"], options["usd_rate"], options["markup"])
                wholesale_price = self.calculate_wholesale_price(stock_row["usd"], options["usd_rate"])

                color_name, color_hex = item["color"]
                color, _ = ProductColor.objects.get_or_create(name=color_name, defaults={"hash_code": color_hex})
                if color.hash_code != color_hex:
                    color.hash_code = color_hex
                    color.save(update_fields=["hash_code"])

                memory = None
                if item.get("storage"):
                    memory, _ = Memory.objects.get_or_create(volume=item["storage"])

                size = None
                if item.get("size"):
                    size, _ = ProductSize.objects.get_or_create(name=item["size"])

                product = self.upsert_product(item, category, brand, brand_category, color, memory, size)
                variant = self.upsert_variant(item, product, color, memory, size)
                self.upsert_stock(variant, channel.shop, stock_row["quantity"], wholesale_price)
                price_obj = self.upsert_price(variant, channel, price)
                payload_ids.append(price_obj.id)
                if not options["skip_images"]:
                    self.ensure_images(variant, item)

        adapter = get_marketplace_adapter(channel)
        payload = adapter.build_payload(channel_price_ids=payload_ids)
        self.stdout.write(f"Подготовлено товаров для M-Market: {len(payload['products'])}")
        self.stdout.write(f"ChannelPrice IDs: {payload_ids}")
        self.stdout.write(json.dumps(payload, ensure_ascii=False, indent=2))

        if options["dry_run"]:
            self.stdout.write(self.style.WARNING("Dry-run: товары созданы/обновлены, но запрос в M-Market не отправлен."))
            return

        result = adapter.push_products(channel_price_ids=payload_ids)
        self.stdout.write(self.style.SUCCESS(f"Отправлено в {channel}: {result}"))

    def get_channel(self, options):
        if options["channel_id"]:
            channel = Channel.objects.filter(id=options["channel_id"]).first()
        else:
            channel = Channel.objects.filter(adapter_key=options["adapter_key"], is_active=True).first()
        if not channel:
            raise CommandError("Канал не найден. Передай --channel-id или проверь adapter_key.")
        if not channel.api_url or not channel.api_token or not channel.branch_id:
            raise CommandError(f"У канала {channel} не заполнены API URL/API token/branch_id.")
        return channel

    def read_ipads(self, path):
        workbook = load_workbook(self.resolve_excel_path(path), data_only=True)
        sheet = workbook.active
        rows = OrderedDict()
        for row in sheet.iter_rows(values_only=True):
            name = row[0]
            if not name or "ipad" not in str(name).casefold():
                continue
            key = self.normalize_excel_name(name)
            quantity = self.decimal_value(row[1]) or Decimal("0")
            usd = self.decimal_value(row[2])
            if not usd:
                continue
            entry = rows.setdefault(key, {"quantity": Decimal("0"), "usd": usd})
            entry["quantity"] += quantity
            entry["usd"] = max(entry["usd"], usd)
        return rows

    def resolve_excel_path(self, path):
        original_path = Path(path)
        candidates = [original_path]
        if not original_path.is_absolute():
            candidates.extend(
                [
                    Path.cwd() / original_path,
                    Path("/app") / original_path,
                    Path("/app/imports") / original_path,
                    Path("/tmp") / original_path,
                ]
            )
        for candidate in candidates:
            if candidate.exists():
                return candidate
        checked = ", ".join(str(candidate) for candidate in candidates)
        raise CommandError(f"Excel-файл не найден. Проверенные пути: {checked}")

    def normalize_excel_name(self, value):
        return " ".join(str(value).upper().replace("Б/У", "B/U").replace("Б.У", "B/U").split())

    def decimal_value(self, value):
        if value is None:
            return None
        if isinstance(value, Decimal):
            return value
        if isinstance(value, (int, float)):
            return Decimal(str(value))
        normalized = str(value).replace(" ", "").replace(",", ".")
        return Decimal(normalized)

    def calculate_price(self, usd_price, usd_rate, markup):
        return (usd_price * usd_rate * (Decimal("1") + markup)).quantize(Decimal("1"), rounding=ROUND_HALF_UP)

    def calculate_wholesale_price(self, usd_price, usd_rate):
        return (usd_price * usd_rate).quantize(Decimal("1"), rounding=ROUND_HALF_UP)

    def build_item(self, key):
        if "MAGIC KEYBOARD" in key:
            return self.build_keyboard_item(key)
        return self.build_ipad_item(key)

    def build_keyboard_item(self, key):
        is_pro = "PRO" in key
        size = "13-inch" if "13" in key or "12.9" in key else "11-inch"
        if is_pro:
            keyboard_type = "Magic Keyboard for iPad Pro"
        elif size == "13-inch":
            keyboard_type = "Magic Keyboard for iPad Air"
        else:
            keyboard_type = "Magic Keyboard Folio for iPad"
        sku = self.extract_model_code(key) or self.sku_from_key(key)
        item = {
            "sku": sku,
            "name": f"{keyboard_type} {size} Black",
            "model": keyboard_type,
            "condition": "Новый",
            "color": COLORS["BLACK"],
            "storage": None,
            "size": size,
            "connectivity": "",
            "chip": "",
            "display": "",
            "camera": "",
            "front_camera": "",
            "is_accessory": True,
            "features": [
                "магнитное крепление к iPad",
                "встроенный трекпад",
                "регулируемый угол наклона",
                "защита спереди и сзади",
            ],
            "image_urls": KEYBOARD_GALLERY["pro" if is_pro else "air" if size == "13-inch" else "folio"],
        }
        item["description"] = (
            f"{item['name']} — оригинальная клавиатура Apple для iPad в цвете Black. "
            f"Подходит для работы, учебы и набора текста, оснащена трекпадом и магнитным креплением. "
            f"Основные особенности: {', '.join(item['features'])}."
        )
        return item

    def build_ipad_item(self, key):
        family = self.detect_family(key)
        storage = self.extract_storage(key)
        size = self.extract_size(key, family)
        connectivity = self.extract_connectivity(key)
        color = self.extract_color(key)
        if family == "iPad Pro" and color[0] == "Black":
            color = COLORS["SPACE BLACK"]
        condition = "Б/У" if "B/U" in key else "Новый"
        chip = self.extract_chip(key, family)
        sku = self.extract_model_code(key) or self.sku_from_key(key)
        model = self.model_name(family, size, chip)
        name_parts = [model]
        if storage:
            name_parts.append(storage)
        if connectivity:
            name_parts.append(connectivity)
        name_parts.append(color[0])
        if condition == "Б/У":
            name_parts.append("B/U")

        display = self.display_for(family, size)
        camera = "12MP Wide camera, 4K video"
        front_camera = "Landscape 12MP Center Stage camera"
        features = [
            display,
            camera,
            front_camera,
            "USB-C",
            "поддержка Apple Pencil",
        ]
        if chip:
            features.insert(1, f"{chip} chip")
        if connectivity == "Wi-Fi + Cellular":
            features.append("5G")

        item = {
            "sku": sku,
            "name": " ".join(name_parts),
            "model": model,
            "condition": condition,
            "color": color,
            "storage": storage,
            "size": size,
            "connectivity": connectivity,
            "chip": chip,
            "display": display,
            "camera": camera,
            "front_camera": front_camera,
            "is_accessory": False,
            "features": features,
            "image_urls": self.image_urls_for(family, size, color[0], connectivity),
        }
        item["description"] = (
            f"{item['name']} — оригинальный планшет Apple, состояние: {condition}. "
            f"Основные характеристики: {', '.join(features)}. "
            f"Память: {storage or 'уточняется'}. Связь: {connectivity}. "
            f"Подходит для работы, учебы, просмотра контента, заметок и повседневных задач."
        )
        return item

    def detect_family(self, key):
        if "IPAD AIR" in key:
            return "iPad Air"
        if "IPAD PRO" in key:
            return "iPad Pro"
        return "iPad"

    def extract_storage(self, key):
        match = re.search(r"(\d{2,4})\s*GB|(?:^|\s)(128|256|512|1000)(?:\s|$)", key)
        if not match:
            return None
        value = match.group(1) or match.group(2)
        return f"{value}GB"

    def extract_size(self, key, family):
        if "12.9" in key:
            return "12.9-inch"
        if re.search(r"\b13\b", key):
            return "13-inch"
        if re.search(r"\b11\b|11\"", key):
            return "11-inch"
        return "11-inch" if family != "iPad Pro" else "13-inch"

    def extract_connectivity(self, key):
        if "CELLULAR" in key or "5G" in key or "WIFI+CELLULAR" in key or "WI-FI CELLULAR" in key:
            return "Wi-Fi + Cellular"
        return "Wi-Fi"

    def extract_color(self, key):
        for marker in ("SPACE BLACK", "S/BLACK", "SPACE GRAY", "SPACEGRAY", "STARLIGHT", "YELLOW", "SILVER", "PINK", "BLUE", "BLACK"):
            if marker in key:
                return COLORS[marker]
        return COLORS["SILVER"]

    def extract_chip(self, key, family):
        for chip in ("M5", "M4", "M3", "M2", "A16"):
            if chip in key:
                return chip
        if "2022" in key and family == "iPad Pro":
            return "M2"
        if family == "iPad":
            return "A16"
        if family == "iPad Air":
            return "M4"
        return None

    def extract_model_code(self, key):
        match = re.search(r"\(([A-Z0-9]+)\)", key)
        if match:
            return f"IPAD-{match.group(1)}"
        slash_match = re.search(r"/(M[A-Z0-9]{3,})", key)
        if slash_match:
            return f"IPAD-{slash_match.group(1)}"
        tail_match = re.search(r"\b(M[A-Z0-9]{3,4})\b$", key)
        if tail_match:
            return f"IPAD-{tail_match.group(1)}"
        return None

    def sku_from_key(self, key):
        cleaned = re.sub(r"[^A-Z0-9]+", "-", key).strip("-")
        digest = hashlib.sha1(key.encode("utf-8")).hexdigest()[:6].upper()
        return f"{cleaned[:70]}-{digest}"

    def model_name(self, family, size, chip):
        if family == "iPad":
            return f"iPad {size} {chip}"
        if chip:
            return f"{family} {size} {chip}"
        return f"{family} {size}"

    def display_for(self, family, size):
        if family == "iPad Pro":
            return f"{size} Ultra Retina XDR OLED display"
        return f"{size} Liquid Retina display"

    def image_urls_for(self, family, size, color_name, connectivity):
        if family == "iPad":
            color_slug = color_name.lower().replace(" ", "")
            connection_slug = "cell" if connectivity == "Wi-Fi + Cellular" else "wifi"
            return [
                store_image(f"ipad-2022-hero-{color_slug}-{connection_slug}-select", 940, 1112, "png-alpha"),
                IPAD_A16_GALLERY[0],
                IPAD_A16_GALLERY[1],
            ]

        if family == "iPad Air":
            color_slug = color_name.lower().replace(" ", "")
            if color_slug == "spacegray":
                color_slug = "spacegray"
            size_slug = "13in" if size == "13-inch" else "11in"
            connection_slug = "cell" if connectivity == "Wi-Fi + Cellular" else "wifi"
            return [
                store_image(f"ipad-air-select-{size_slug}-{connection_slug}-{color_slug}-202405"),
                IPAD_AIR_GALLERY[1],
                IPAD_AIR_GALLERY[2],
            ]

        pro_color = "silver" if color_name == "Silver" else "spaceblack"
        pro_size = "13" if size in {"13-inch", "12.9-inch"} else "11"
        connection_slug = "wificell" if connectivity == "Wi-Fi + Cellular" else "wifi"
        return [
            store_image(f"ipad-pro-{pro_size}-select-{connection_slug}-{pro_color}-202405"),
            IPAD_PRO_GALLERY[0],
            IPAD_PRO_GALLERY[1],
        ]

    def upsert_product(self, item, category, brand, brand_category, color, memory, size):
        product = ProductVariant.objects.filter(sku=item["sku"]).select_related("product").first()
        product = product.product if product else Product.objects.filter(name=item["name"]).first()
        if product is None:
            product = Product(name=item["name"])
        product.category = category.name
        product.category_ref = category
        product.brand_name = brand.name
        product.brand_ref = brand
        product.brand_category = brand_category.name
        product.brand_category_ref = brand_category
        product.description = item["description"][:1000]
        product.save()
        product.colors.add(color)
        if memory:
            product.memories.add(memory)
        if size:
            product.sizes.add(size)
        return product

    def upsert_variant(self, item, product, color, memory, size):
        attrs = {
            "Тип": "Клавиатура" if item["is_accessory"] else "Планшет",
            "Производители": "Apple",
            "Модель": item["model"],
            "Цвет": color.name,
            "Состояние": item["condition"],
        }
        if item["is_accessory"]:
            attrs.update(
                {
                    "Совместимость": item["size"],
                    "Особенности": ", ".join(item["features"]),
                }
            )
        else:
            attrs.update(
                {
                    "Память": item["storage"],
                    "Диагональ экрана": item["size"],
                    "Дисплей": item["display"],
                    "Камера": item["camera"],
                    "Фронтальная камера": item["front_camera"],
                    "Связь": item["connectivity"],
                    "Разъем": "USB-C",
                    "Операционная система": "iPadOS",
                }
            )
            if item["chip"]:
                attrs["Процессор"] = item["chip"]
        variant, _ = ProductVariant.objects.update_or_create(
            sku=item["sku"],
            defaults={
                "product": product,
                "color": color,
                "memory": memory,
                "size": size,
                "attributes": attrs,
                "is_active": True,
            },
        )
        return variant

    def upsert_stock(self, variant, shop, quantity, usd_price):
        Stock.objects.update_or_create(
            variant=variant,
            shop=shop,
            defaults={
                "quantity": int(quantity),
                "in_stock": quantity > 0,
                "wholesale_price": usd_price,
            },
        )

    def upsert_price(self, variant, channel, price):
        price_obj, _ = ChannelPrice.objects.update_or_create(
            variant=variant,
            shop=channel.shop,
            channel=channel,
            defaults={
                "price": price,
                "sync_status": ChannelPrice.SyncStatus.PENDING,
                "last_sync_error": "",
            },
        )
        return price_obj

    def ensure_images(self, variant, item):
        existing_images = list(variant.images.all())
        has_valid_images = len(existing_images) >= 3 and all(
            urlparse(image.image.url).path.lower().endswith((".jpg", ".jpeg", ".png", ".webp"))
            for image in existing_images
        )
        if has_valid_images:
            return
        variant.images.all().delete()
        for index, image_url in enumerate(item["image_urls"]):
            content = self.download_image(image_url, f"{item['sku'].lower()}_{index}.jpg")
            ProductImage.objects.create(
                variant=variant,
                image=content,
                color=variant.color,
                is_primary=index == 0,
                order=index,
            )

    def download_image(self, url, filename):
        try:
            response = requests.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=30)
            response.raise_for_status()
            image = Image.open(BytesIO(response.content)).convert("RGBA")
        except (requests.RequestException, OSError) as exc:
            raise CommandError(f"Фото {url} недоступно: {exc}") from exc

        canvas = Image.new("RGBA", image.size, "WHITE")
        canvas.alpha_composite(image)
        rgb_image = canvas.convert("RGB")
        output = BytesIO()
        rgb_image.save(output, format="JPEG", quality=92, optimize=True)
        return ContentFile(output.getvalue(), name=filename)
