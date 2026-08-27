from collections import OrderedDict
from decimal import Decimal, ROUND_HALF_UP
import json
from pathlib import PurePosixPath
from urllib.parse import urlparse

import requests
from django.core.files.base import ContentFile
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from core.marketplace.factory import get_marketplace_adapter
from core.models import (
    Brand,
    BrandCategory,
    Channel,
    ChannelPrice,
    Product,
    ProductCategory,
    ProductColor,
    ProductImage,
    ProductVariant,
    Stock,
)


APPLE_BASE = "https://www.apple.com"


PRODUCTS = {
    "AIRPODS 4 - ACTIVE NOISE CANCELLATION (MXP93)": {
        "sku": "AP4ANC-MXP93",
        "name": "AirPods 4 with Active Noise Cancellation",
        "model": "AirPods 4 ANC",
        "condition": "Новый",
        "color": ("White", "#F5F5F7"),
        "features": [
            "Active Noise Cancellation",
            "Adaptive Audio",
            "Transparency mode",
            "Conversation Awareness",
            "Personalized Spatial Audio with dynamic head tracking",
            "Dust, sweat, and water resistant IP54",
        ],
        "battery": "Up to 4 hours listening time with ANC on; up to 20 hours with case and ANC on",
        "dimensions": "AirPods: 30.2 x 18.3 x 18.1 mm; charging case: 46.2 x 50.1 x 21.2 mm",
        "case": "Charging Case (USB-C) with speaker",
        "image_urls": [
            f"{APPLE_BASE}/v/airpods-4/g/images/specs/hero_airpods_4gen_active__bofs1xp3rnaq_large.jpg",
            f"{APPLE_BASE}/v/airpods-4/g/images/specs/airpods_height_width__dvf3datp2auu_large.jpg",
            f"{APPLE_BASE}/v/airpods-4/g/images/specs/case_width__dcgzagkt9bee_large.jpg",
        ],
    },
    "AIRPODS 4 (MXP63)": {
        "sku": "AP4-MXP63",
        "name": "AirPods 4",
        "model": "AirPods 4",
        "condition": "Новый",
        "color": ("White", "#F5F5F7"),
        "features": [
            "H2 headphone chip",
            "Voice Isolation",
            "Personalized Spatial Audio with dynamic head tracking",
            "Force sensor controls",
            "Dust, sweat, and water resistant IP54",
        ],
        "battery": "Up to 5 hours listening time; up to 30 hours with Charging Case (USB-C)",
        "dimensions": "AirPods: 30.2 x 18.3 x 18.1 mm; charging case: 46.2 x 50.1 x 21.2 mm",
        "case": "Charging Case (USB-C)",
        "image_urls": [
            f"{APPLE_BASE}/v/airpods-4/g/images/specs/hero_airpods_4gen__e53e9vzroy2q_large.jpg",
            f"{APPLE_BASE}/v/airpods-4/g/images/specs/airpods_height_width__dvf3datp2auu_large.jpg",
            f"{APPLE_BASE}/v/airpods-4/g/images/specs/case_height__fx3ygamudsuy_large.jpg",
        ],
    },
    "AIRPODS PRO 3 (MFHP4)": {
        "sku": "APPRO3-MFHP4",
        "name": "AirPods Pro 3",
        "model": "AirPods Pro 3",
        "condition": "Новый",
        "color": ("White", "#F5F5F7"),
        "features": [
            "Active Noise Cancellation",
            "Adaptive Audio",
            "Transparency mode",
            "Heart rate sensor for workouts",
            "Apple H2 headphone chip",
            "Dust, sweat, and water resistant IP57",
        ],
        "battery": "Up to 8 hours listening time with ANC; up to 24 hours with MagSafe Charging Case (USB-C)",
        "dimensions": "AirPods Pro 3: 30.9 x 19.2 x 27.0 mm; case: 47.2 x 62.2 x 21.8 mm",
        "case": "MagSafe Charging Case (USB-C) with speaker and lanyard loop",
        "image_urls": [
            f"{APPLE_BASE}/v/airpods-pro/s/images/specs/airpods__eqrzs6rwhu2q_large.jpg",
            f"{APPLE_BASE}/v/airpods-pro/s/images/specs/airpods_width_height__b4ym4bd8ahau_large.jpg",
            f"{APPLE_BASE}/v/airpods-pro/s/images/specs/case_width_height__9r5ux142uqqi_large.jpg",
        ],
    },
    "AIRPODS PRO 3": {
        "sku": "APPRO3",
        "name": "AirPods Pro 3",
        "model": "AirPods Pro 3",
        "condition": "Новый",
        "color": ("White", "#F5F5F7"),
        "features": [
            "Active Noise Cancellation",
            "Adaptive Audio",
            "Transparency mode",
            "Heart rate sensor for workouts",
            "Apple H2 headphone chip",
            "Dust, sweat, and water resistant IP57",
        ],
        "battery": "Up to 8 hours listening time with ANC; up to 24 hours with MagSafe Charging Case (USB-C)",
        "dimensions": "AirPods Pro 3: 30.9 x 19.2 x 27.0 mm; case: 47.2 x 62.2 x 21.8 mm",
        "case": "MagSafe Charging Case (USB-C) with speaker and lanyard loop",
        "image_urls": [
            f"{APPLE_BASE}/v/airpods-pro/s/images/specs/airpods__eqrzs6rwhu2q_large.jpg",
            f"{APPLE_BASE}/v/airpods-pro/s/images/specs/airpods_width_height__b4ym4bd8ahau_large.jpg",
            f"{APPLE_BASE}/v/airpods-pro/s/images/specs/case_depth__l2ozvqw3msae_large.jpg",
        ],
    },
    "AIRPODS 4 ANC B/U": {
        "sku": "AP4ANC-BU",
        "name": "AirPods 4 with Active Noise Cancellation B/U",
        "model": "AirPods 4 ANC",
        "condition": "Б/У",
        "color": ("White", "#F5F5F7"),
        "features": [
            "Active Noise Cancellation",
            "Adaptive Audio",
            "Transparency mode",
            "Conversation Awareness",
            "Personalized Spatial Audio with dynamic head tracking",
            "Dust, sweat, and water resistant IP54",
        ],
        "battery": "Up to 4 hours listening time with ANC on; up to 20 hours with case and ANC on",
        "dimensions": "AirPods: 30.2 x 18.3 x 18.1 mm; charging case: 46.2 x 50.1 x 21.2 mm",
        "case": "Charging Case (USB-C) with speaker",
        "image_urls": [
            f"{APPLE_BASE}/v/airpods-4/g/images/specs/hero_airpods_4gen_active__bofs1xp3rnaq_large.jpg",
            f"{APPLE_BASE}/v/airpods-4/g/images/specs/airpods_height_width__dvf3datp2auu_large.jpg",
            f"{APPLE_BASE}/v/airpods-4/g/images/specs/case_width__dcgzagkt9bee_large.jpg",
        ],
    },
}


MAX_COLOR_IMAGES = {
    "AIRPODS MAX 2 (USB-C) - PURPLE (MHWP4)": ("APMAX2-MHWP4", "Purple", "#C8B5D9", "hero_purple__skk4zesfid6y"),
    "AIRPODS MAX 2 (USB-C) BLUE (MHWM4)": ("APMAX2-MHWM4", "Blue", "#9FC4D2", "hero_blue__f6wblks3x02m"),
    "AIRPODS MAX 2 (USB-C) MIDNIGHT(MHWK4)": ("APMAX2-MHWK4", "Midnight", "#1F2933", "hero_midnight__d5qsulrfqzu6"),
    "AIRPODS MAX 2 (USB-C) ORANGE (MHWN4)": ("APMAX2-MHWN4", "Orange", "#F2A07B", "hero_orange__dujzhx2hoo8y"),
    "AIRPODS MAX 2 (USB-C) STARLIGHT (MHWL4)": ("APMAX2-MHWL4", "Starlight", "#E6DCCD", "hero_starlight__yc0aejmzknm6"),
}


for excel_name, (sku, color_name, color_hex, hero) in MAX_COLOR_IMAGES.items():
    PRODUCTS[excel_name] = {
        "sku": sku,
        "name": f"AirPods Max 2 USB-C {color_name}",
        "model": "AirPods Max 2",
        "condition": "Новый",
        "color": (color_name, color_hex),
        "features": [
            "Active Noise Cancellation",
            "Adaptive Audio",
            "Transparency mode",
            "Personalized Spatial Audio with dynamic head tracking",
            "Apple H2 headphone chip in each ear cup",
            "Lossless Audio and ultra-low latency audio via USB-C",
        ],
        "battery": "Up to 20 hours listening time with Active Noise Cancellation enabled",
        "dimensions": "AirPods Max 2 including cushions: 168.6 x 187.3 x 83.4 mm; weight 386.2 g",
        "case": "Smart Case and USB-C Charge Cable",
        "image_urls": [
            f"{APPLE_BASE}/v/airpods-max/k/images/specs/{hero}_large.jpg",
            f"{APPLE_BASE}/v/airpods-max/k/images/specs/airpods_front__bwbxvilb50eq_large.jpg",
            f"{APPLE_BASE}/v/airpods-max/k/images/specs/airpods_side__fsa2bfynojmi_large.jpg",
        ],
    }


class Command(BaseCommand):
    help = "Загружает AirPods из Excel Turan в канал SHAT / M-Market."

    def add_arguments(self, parser):
        parser.add_argument(
            "--excel",
            default="/Users/islamaibekov/Downloads/Price Turan .xlsx",
            help="Путь к Excel-файлу с остатками Turan.",
        )
        parser.add_argument("--channel-id", type=int, help="ID канала M-Market.")
        parser.add_argument("--adapter-key", default="mmarket", help="Ключ адаптера канала.")
        parser.add_argument("--markup", type=Decimal, default=Decimal("0.15"), help="Наценка, например 0.15.")
        parser.add_argument("--usd-rate", type=Decimal, default=Decimal("88"), help="Курс USD/KGS.")
        parser.add_argument("--dry-run", action="store_true", help="Создать товары и показать payload без отправки.")
        parser.add_argument("--skip-images", action="store_true", help="Не скачивать фото, если их уже минимум 3.")

    def handle(self, *args, **options):
        channel = self.get_channel(options)
        rows = self.read_airpods(options["excel"])
        payload_ids = []

        with transaction.atomic():
            brand, _ = Brand.objects.get_or_create(name="Apple")
            brand_category, _ = BrandCategory.objects.get_or_create(brand=brand, name="AirPods")
            category, _ = ProductCategory.objects.get_or_create(name="Наушники")

            for key, stock_row in rows.items():
                item = PRODUCTS.get(key)
                if not item:
                    self.stdout.write(self.style.WARNING(f"Пропущено, нет описания: {key}"))
                    continue
                price = self.calculate_price(stock_row["usd"], options["usd_rate"], options["markup"])
                color_name, color_hex = item["color"]
                color, _ = ProductColor.objects.get_or_create(name=color_name, defaults={"hash_code": color_hex})
                if color.hash_code != color_hex:
                    color.hash_code = color_hex
                    color.save(update_fields=["hash_code"])

                product = self.upsert_product(item, category, brand, brand_category, color)
                variant = self.upsert_variant(item, product, color)
                wholesale_price = self.calculate_wholesale_price(stock_row["usd"], options["usd_rate"])
                self.upsert_stock(variant, channel.shop, stock_row["quantity"], wholesale_price)
                price_obj = self.upsert_price(variant, channel, price)
                payload_ids.append(price_obj.id)
                if not options["skip_images"]:
                    self.ensure_images(variant, item)

        adapter = get_marketplace_adapter(channel)
        payload = adapter.build_payload(channel_price_ids=payload_ids)
        self.stdout.write(f"Подготовлено товаров для M-Market: {len(payload['products'])}")
        self.stdout.write(f"ChannelPrice IDs: {payload_ids}")
        self.stdout.write(json.dumps(payload, ensure_ascii=False, indent=2))

        if options["dry_run"]:
            self.stdout.write(self.style.WARNING("Dry-run: товары созданы/обновлены, но запрос в M-Market не отправлен."))
            return

        result = adapter.push_products(channel_price_ids=payload_ids)
        self.stdout.write(self.style.SUCCESS(f"Отправлено в {channel}: {result}"))

    def get_channel(self, options):
        if options["channel_id"]:
            channel = Channel.objects.filter(id=options["channel_id"]).first()
        else:
            channel = Channel.objects.filter(adapter_key=options["adapter_key"], is_active=True).first()
        if not channel:
            raise CommandError("Канал не найден. Передай --channel-id или проверь adapter_key.")
        if not channel.api_url or not channel.api_token or not channel.branch_id:
            raise CommandError(f"У канала {channel} не заполнены API URL/API token/branch_id.")
        return channel

    def read_airpods(self, path):
        workbook = load_workbook(path, data_only=True)
        sheet = workbook.active
        rows = OrderedDict()
        for row in sheet.iter_rows(values_only=True):
            name = row[0]
            if not name or "airpods" not in str(name).casefold():
                continue
            key = self.normalize_excel_name(name)
            quantity = self.decimal_value(row[1]) or Decimal("0")
            usd = self.decimal_value(row[2])
            if not usd:
                continue
            entry = rows.setdefault(key, {"quantity": Decimal("0"), "usd": usd})
            entry["quantity"] += quantity
            entry["usd"] = max(entry["usd"], usd)
        return rows

    def normalize_excel_name(self, value):
        return " ".join(str(value).upper().replace("Б/У", "B/U").replace("Б.У", "B/U").split())

    def decimal_value(self, value):
        if value is None:
            return None
        if isinstance(value, Decimal):
            return value
        if isinstance(value, (int, float)):
            return Decimal(str(value))
        normalized = str(value).replace(" ", "").replace(",", ".")
        return Decimal(normalized)

    def calculate_price(self, usd_price, usd_rate, markup):
        return (usd_price * usd_rate * (Decimal("1") + markup)).quantize(Decimal("1"), rounding=ROUND_HALF_UP)

    def calculate_wholesale_price(self, usd_price, usd_rate):
        return (usd_price * usd_rate).quantize(Decimal("1"), rounding=ROUND_HALF_UP)

    def upsert_product(self, item, category, brand, brand_category, color):
        description = (
            f"{item['name']} — оригинальные наушники Apple, состояние: {item['condition']}. "
            f"Основные возможности: {', '.join(item['features'])}. "
            f"Автономность: {item['battery']}. Размеры: {item['dimensions']}. "
            f"Комплект: {item['case']}. Подходят для iPhone, iPad, Mac, Apple Watch и Bluetooth-устройств."
        )
        product = ProductVariant.objects.filter(sku=item["sku"]).select_related("product").first()
        product = product.product if product else Product.objects.filter(name=item["name"]).first()
        if product is None:
            product = Product(name=item["name"])
        product.category = category.name
        product.category_ref = category
        product.brand_name = brand.name
        product.brand_ref = brand
        product.brand_category = brand_category.name
        product.brand_category_ref = brand_category
        product.description = description[:1000]
        product.save()
        product.colors.add(color)
        return product

    def upsert_variant(self, item, product, color):
        attrs = {
            "Тип": "AirPods",
            "Производители": "Apple",
            "Модель": item["model"],
            "Цвет": color.name,
            "Состояние": item["condition"],
            "Чип": "Apple H2",
            "Беспроводная связь": "Bluetooth 5.3",
            "Автономность": item["battery"],
            "Комплект": item["case"],
        }
        variant, _ = ProductVariant.objects.update_or_create(
            sku=item["sku"],
            defaults={
                "product": product,
                "color": color,
                "attributes": attrs,
                "is_active": True,
            },
        )
        return variant

    def upsert_stock(self, variant, shop, quantity, usd_price):
        Stock.objects.update_or_create(
            variant=variant,
            shop=shop,
            defaults={
                "quantity": int(quantity),
                "in_stock": quantity > 0,
                "wholesale_price": usd_price,
            },
        )

    def upsert_price(self, variant, channel, price):
        price_obj, _ = ChannelPrice.objects.update_or_create(
            variant=variant,
            shop=channel.shop,
            channel=channel,
            defaults={
                "price": price,
                "sync_status": ChannelPrice.SyncStatus.PENDING,
                "last_sync_error": "",
            },
        )
        return price_obj

    def ensure_images(self, variant, item):
        existing_images = list(variant.images.all())
        has_valid_images = len(existing_images) >= 3 and all(
            urlparse(image.image.url).path.lower().endswith((".jpg", ".jpeg", ".png", ".webp"))
            for image in existing_images
        )
        if has_valid_images:
            return
        variant.images.all().delete()
        for index, image_url in enumerate(item["image_urls"]):
            suffix = PurePosixPath(urlparse(image_url).path).suffix.lower()
            if suffix not in {".jpg", ".jpeg", ".png", ".webp"}:
                suffix = ".jpg"
            content = self.download_image(image_url, f"{item['sku'].lower()}_{index}{suffix}")
            ProductImage.objects.create(
                variant=variant,
                image=content,
                color=variant.color,
                is_primary=index == 0,
                order=index,
            )

    def download_image(self, url, filename):
        try:
            response = requests.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=30)
            response.raise_for_status()
        except requests.RequestException as exc:
            raise CommandError(f"Фото {url} недоступно: {exc}") from exc
        return ContentFile(response.content, name=filename)
