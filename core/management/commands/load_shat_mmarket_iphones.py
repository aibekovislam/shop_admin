from collections import OrderedDict
from decimal import Decimal, ROUND_HALF_UP
from io import BytesIO
import hashlib
import json
import re
from pathlib import Path
from urllib.parse import urlparse

from django.core.files.base import ContentFile
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook
from PIL import Image
import requests

from core.marketplace.factory import get_marketplace_adapter
from core.models import (
    Brand,
    BrandCategory,
    Channel,
    ChannelPrice,
    Memory,
    Product,
    ProductCategory,
    ProductColor,
    ProductImage,
    ProductVariant,
    Stock,
)


STORE_IMAGE_BASE = "https://store.storeimages.cdn-apple.com/1/as-images.apple.com/is/"


COLORS = {
    "BLACK": ("Black", "#1D1D1F"),
    "JET BLACK": ("Jet Black", "#101010"),
    "WHITE": ("White", "#F5F5F7"),
    "PINK": ("Pink", "#F7B6CF"),
    "TEAL": ("Teal", "#9AD7D2"),
    "ULTRAMARINE": ("Ultramarine", "#5E6FEF"),
    "BLUE": ("Blue", "#8FB9E8"),
    "DEEP BLUE": ("Deep Blue", "#1F314F"),
    "SKY BLUE": ("Sky Blue", "#C8DDEE"),
    "MIST BLUE": ("Mist Blue", "#C8DDEE"),
    "GREEN": ("Green", "#BFD7B5"),
    "SAGE": ("Sage", "#BFD7B5"),
    "PURPLE": ("Deep Purple", "#594F63"),
    "DEEP PURPLE": ("Deep Purple", "#594F63"),
    "SILVER": ("Silver", "#E3E4E5"),
    "ORANGE": ("Cosmic Orange", "#F77F45"),
    "COSMIC ORANGE": ("Cosmic Orange", "#F77F45"),
    "DESERT": ("Desert Titanium", "#CDB9A5"),
    "DESERT TITANIUM": ("Desert Titanium", "#CDB9A5"),
    "NATURAL": ("Natural Titanium", "#B9B5AC"),
    "NATURAL TITANIUM": ("Natural Titanium", "#B9B5AC"),
    "SPACE BLACK": ("Space Black", "#2E2E30"),
}


EXCLUDED_IPHONE_WORDS = (
    "АДАПТЕР",
    "ADAPTER",
    "ЧЕХОЛ",
    "CASE",
    "CABLE",
    "HOCO",
)


REGION_MARKERS = (
    "LLA",
    "JA",
    "JAPAN",
    "HK",
    "AU",
    "KR",
    "EAC",
    "LATIN",
    "SPC",
    "ACTIVE",
    "ASIS",
    "2SIM",
    "E SIM",
    "ESIM",
)


def store_image(image_id, width=900, height=900, image_format="jpeg"):
    return f"{STORE_IMAGE_BASE}{image_id}?wid={width}&hei={height}&fmt={image_format}"


class Command(BaseCommand):
    help = "Загружает iPhone из Excel Turan в канал SHAT / M-Market."

    def add_arguments(self, parser):
        parser.add_argument(
            "--excel",
            default="Price Turan .xlsx",
            help="Путь к Excel-файлу с остатками Turan.",
        )
        parser.add_argument("--channel-id", type=int, help="ID канала M-Market.")
        parser.add_argument("--adapter-key", default="mmarket", help="Ключ адаптера канала.")
        parser.add_argument("--markup", type=Decimal, default=Decimal("0.15"), help="Наценка, например 0.15.")
        parser.add_argument("--usd-rate", type=Decimal, default=Decimal("88"), help="Курс USD/KGS.")
        parser.add_argument("--dry-run", action="store_true", help="Создать товары и показать payload без отправки.")
        parser.add_argument("--skip-images", action="store_true", help="Не скачивать фото, если их уже минимум 3.")

    def handle(self, *args, **options):
        channel = self.get_channel(options)
        rows = self.read_iphones(options["excel"])
        payload_ids = []

        with transaction.atomic():
            brand, _ = Brand.objects.get_or_create(name="Apple")
            brand_category, _ = BrandCategory.objects.get_or_create(brand=brand, name="iPhone")
            category, _ = ProductCategory.objects.get_or_create(name="Смартфоны")

            for key, stock_row in rows.items():
                item = self.build_item(key)
                price = self.calculate_price(stock_row["usd"], options["usd_rate"], options["markup"])
                wholesale_price = self.calculate_wholesale_price(stock_row["usd"], options["usd_rate"])

                color_name, color_hex = item["color"]
                color, _ = ProductColor.objects.get_or_create(name=color_name, defaults={"hash_code": color_hex})
                if color.hash_code != color_hex:
                    color.hash_code = color_hex
                    color.save(update_fields=["hash_code"])

                memory, _ = Memory.objects.get_or_create(volume=item["storage"])
                product = self.upsert_product(item, category, brand, brand_category, color, memory)
                variant = self.upsert_variant(item, product, color, memory)
                self.upsert_stock(variant, channel.shop, stock_row["quantity"], wholesale_price)
                price_obj = self.upsert_price(variant, channel, price)
                payload_ids.append(price_obj.id)
                if not options["skip_images"]:
                    self.ensure_images(variant, item)

        adapter = get_marketplace_adapter(channel)
        payload = adapter.build_payload(channel_price_ids=payload_ids)
        self.stdout.write(f"Подготовлено товаров для M-Market: {len(payload['products'])}")
        self.stdout.write(f"ChannelPrice IDs: {payload_ids}")
        self.stdout.write(json.dumps(payload, ensure_ascii=False, indent=2))

        if options["dry_run"]:
            self.stdout.write(self.style.WARNING("Dry-run: товары созданы/обновлены, но запрос в M-Market не отправлен."))
            return

        result = adapter.push_products(channel_price_ids=payload_ids)
        self.stdout.write(self.style.SUCCESS(f"Отправлено в {channel}: {result}"))

    def get_channel(self, options):
        if options["channel_id"]:
            channel = Channel.objects.filter(id=options["channel_id"]).first()
        else:
            channel = Channel.objects.filter(adapter_key=options["adapter_key"], is_active=True).first()
        if not channel:
            raise CommandError("Канал не найден. Передай --channel-id или проверь adapter_key.")
        if not channel.api_url or not channel.api_token or not channel.branch_id:
            raise CommandError(f"У канала {channel} не заполнены API URL/API token/branch_id.")
        return channel

    def read_iphones(self, path):
        workbook = load_workbook(self.resolve_excel_path(path), data_only=True)
        sheet = workbook.active
        rows = OrderedDict()
        for row in sheet.iter_rows(values_only=True):
            name = row[0]
            if not name:
                continue
            key = self.normalize_excel_name(name)
            if not self.is_iphone_row(key):
                continue
            quantity = self.decimal_value(row[1]) or Decimal("0")
            usd = self.decimal_value(row[2])
            if not usd:
                continue
            entry = rows.setdefault(key, {"quantity": Decimal("0"), "usd": usd})
            entry["quantity"] += quantity
            entry["usd"] = max(entry["usd"], usd)
        return rows

    def is_iphone_row(self, key):
        if any(word in key for word in EXCLUDED_IPHONE_WORDS):
            return False
        if "IPHONE" in key:
            return True
        if "42MM" in key or "44MM" in key or "46MM" in key:
            return False
        return bool(re.search(r"^(13|14|15|16|17)\b", key) and re.search(r"(GB|TB|PRO|MAX|AIR)", key))

    def resolve_excel_path(self, path):
        original_path = Path(path)
        candidates = [original_path]
        if not original_path.is_absolute():
            candidates.extend(
                [
                    Path.cwd() / original_path,
                    Path("/app") / original_path,
                    Path("/app/imports") / original_path,
                    Path("/tmp") / original_path,
                ]
            )
        for candidate in candidates:
            if candidate.exists():
                return candidate
        checked = ", ".join(str(candidate) for candidate in candidates)
        raise CommandError(f"Excel-файл не найден. Проверенные пути: {checked}")

    def normalize_excel_name(self, value):
        return " ".join(
            str(value)
            .upper()
            .replace("Б/У", "B/U")
            .replace("Б.У", "B/U")
            .replace("Б\\У", "B/U")
            .replace("B/U.", "B/U")
            .split()
        )

    def decimal_value(self, value):
        if value is None:
            return None
        if isinstance(value, Decimal):
            return value
        if isinstance(value, (int, float)):
            return Decimal(str(value))
        normalized = str(value).replace(" ", "").replace(",", ".")
        return Decimal(normalized)

    def calculate_price(self, usd_price, usd_rate, markup):
        return (usd_price * usd_rate * (Decimal("1") + markup)).quantize(Decimal("1"), rounding=ROUND_HALF_UP)

    def calculate_wholesale_price(self, usd_price, usd_rate):
        return (usd_price * usd_rate).quantize(Decimal("1"), rounding=ROUND_HALF_UP)

    def build_item(self, key):
        normalized_key = key if key.startswith("IPHONE") else f"IPHONE {key}"
        model = self.extract_model(normalized_key)
        storage = self.extract_storage(normalized_key)
        color = self.extract_color(normalized_key, model)
        condition = "Б/У" if "B/U" in normalized_key else "Новый"
        region = self.extract_region(normalized_key)
        specs = self.specs_for(model)
        sku = self.sku_from_key(normalized_key)

        name_parts = [model, storage, color[0]]
        if region:
            name_parts.append(region)
        if condition == "Б/У":
            name_parts.append("B/U")

        features = [
            specs["screen"],
            f"{specs['chip']} chip",
            specs["camera"],
            specs["front_camera"],
            "5G",
            "USB-C" if specs["connector"] == "USB-C" else "Lightning",
        ]
        description = (
            f"{' '.join(name_parts)} — оригинальный смартфон Apple, состояние: {condition}. "
            f"Основные характеристики: {', '.join(features)}. "
            f"Память: {storage}. Цвет: {color[0]}. "
            "Подходит для звонков, мессенджеров, фото, видео, игр и ежедневной работы."
        )
        return {
            "sku": sku,
            "name": " ".join(name_parts),
            "model": model,
            "storage": storage,
            "color": color,
            "condition": condition,
            "region": region,
            "description": description[:1000],
            "specs": specs,
            "image_urls": self.image_urls_for(model, color[0]),
        }

    def extract_model(self, key):
        if "17 AIR" in key:
            return "iPhone Air"
        for version in ("17", "16", "15", "14", "13"):
            if f"IPHONE {version}E" in key:
                return f"iPhone {version}e"
            if f"IPHONE {version} PRO MAX" in key:
                return f"iPhone {version} Pro Max"
            if f"IPHONE {version} PRO" in key:
                return f"iPhone {version} Pro"
            if f"IPHONE {version} PLUS" in key:
                return f"iPhone {version} Plus"
            if re.search(rf"\bIPHONE {version}\b", key):
                return f"iPhone {version}"
        raise CommandError(f"Не удалось определить модель iPhone: {key}")

    def extract_storage(self, key):
        tb_match = re.search(r"(\d+)\s*TB", key)
        if tb_match:
            return f"{tb_match.group(1)}TB"
        gb_match = re.search(r"(\d{2,4})\s*GB|(?:^|\s)(128|256|512)(?:\s|$)", key)
        if not gb_match:
            raise CommandError(f"Не удалось определить память iPhone: {key}")
        return f"{gb_match.group(1) or gb_match.group(2)}GB"

    def extract_color(self, key, model):
        if "ORANGE" in key and "17 PRO" in key:
            return COLORS["COSMIC ORANGE"]
        if "BLUE" in key and "17 PRO" in key:
            return COLORS["DEEP BLUE"]
        if "BLUE" in key and model == "iPhone Air":
            return COLORS["SKY BLUE"]
        if "BLACK" in key and model == "iPhone Air":
            return COLORS["SPACE BLACK"]
        if "WHITE" in key and model in {"iPhone 15 Pro", "iPhone 15 Pro Max", "iPhone 16 Pro", "iPhone 16 Pro Max"}:
            return ("White Titanium", "#F5F2EC")
        if "BLUE" in key and model in {"iPhone 15 Pro", "iPhone 15 Pro Max"}:
            return ("Blue Titanium", "#4E5968")
        for marker in (
            "DEEP PURPLE",
            "JET BLACK",
            "SPACE BLACK",
            "SKY BLUE",
            "MIST BLUE",
            "ULTRAMARINE",
            "DESERT TITANIUM",
            "NATURAL TITANIUM",
            "PURPLE",
            "SILVER",
            "BLACK",
            "WHITE",
            "PINK",
            "TEAL",
            "GREEN",
            "SAGE",
            "BLUE",
            "DESERT",
            "NATURAL",
            "ORANGE",
        ):
            if marker in key:
                return COLORS[marker]
        return COLORS["BLACK"]

    def extract_region(self, key):
        found = []
        compact = key.replace("(E SIM)", "E SIM").replace("(ESIM)", "ESIM")
        for marker in REGION_MARKERS:
            if re.search(rf"(^|\s|\(|/){re.escape(marker)}($|\s|\))", compact):
                found.append(marker.replace("E SIM", "eSIM").replace("ESIM", "eSIM"))
        return " ".join(dict.fromkeys(found))

    def specs_for(self, model):
        version_match = re.search(r"iPhone (\d+)", model)
        version = int(version_match.group(1)) if version_match else 17
        is_pro = "Pro" in model
        is_max = "Max" in model
        is_plus = "Plus" in model

        if model == "iPhone Air":
            return {
                "screen": "6.5-inch Super Retina XDR OLED display",
                "screen_size": "6.5",
                "chip": "A19 Pro",
                "camera": "48MP Fusion Main camera",
                "front_camera": "18MP Center Stage front camera",
                "battery": "Video playback up to 27 hours",
                "connector": "USB-C",
                "network": "5G",
            }
        if version >= 17 and is_pro:
            return {
                "screen": ("6.9-inch" if is_max else "6.3-inch") + " Super Retina XDR OLED display with ProMotion",
                "screen_size": "6.9" if is_max else "6.3",
                "chip": "A19 Pro",
                "camera": "48MP Pro Fusion camera system",
                "front_camera": "18MP Center Stage front camera",
                "battery": "Video playback up to 39 hours" if is_max else "Video playback up to 33 hours",
                "connector": "USB-C",
                "network": "5G",
            }
        if version >= 17:
            return {
                "screen": "6.3-inch Super Retina XDR OLED display with ProMotion",
                "screen_size": "6.3",
                "chip": "A19",
                "camera": "48MP Dual Fusion camera system",
                "front_camera": "18MP Center Stage front camera",
                "battery": "Video playback up to 30 hours",
                "connector": "USB-C",
                "network": "5G",
            }
        if model == "iPhone 16e":
            return {
                "screen": "6.1-inch Super Retina XDR OLED display",
                "screen_size": "6.1",
                "chip": "A18",
                "camera": "48MP Fusion camera system",
                "front_camera": "12MP TrueDepth front camera",
                "battery": "Video playback up to 26 hours",
                "connector": "USB-C",
                "network": "5G",
            }
        if version == 16 and is_pro:
            return {
                "screen": ("6.9-inch" if is_max else "6.3-inch") + " Super Retina XDR OLED display with ProMotion",
                "screen_size": "6.9" if is_max else "6.3",
                "chip": "A18 Pro",
                "camera": "48MP Fusion Main + 48MP Ultra Wide + 12MP Telephoto",
                "front_camera": "12MP TrueDepth front camera",
                "battery": "Video playback up to 33 hours" if is_max else "Video playback up to 27 hours",
                "connector": "USB-C",
                "network": "5G",
            }
        if version == 16:
            return {
                "screen": ("6.7-inch" if is_plus else "6.1-inch") + " Super Retina XDR OLED display",
                "screen_size": "6.7" if is_plus else "6.1",
                "chip": "A18",
                "camera": "48MP Fusion Main + 12MP Ultra Wide",
                "front_camera": "12MP TrueDepth front camera",
                "battery": "Video playback up to 27 hours" if is_plus else "Video playback up to 22 hours",
                "connector": "USB-C",
                "network": "5G",
            }
        if version == 15 and is_pro:
            return {
                "screen": ("6.7-inch" if is_max else "6.1-inch") + " Super Retina XDR OLED display with ProMotion",
                "screen_size": "6.7" if is_max else "6.1",
                "chip": "A17 Pro",
                "camera": "48MP Main + Ultra Wide + Telephoto",
                "front_camera": "12MP TrueDepth front camera",
                "battery": "Video playback up to 29 hours" if is_max else "Video playback up to 23 hours",
                "connector": "USB-C",
                "network": "5G",
            }
        if version == 15:
            return {
                "screen": "6.1-inch Super Retina XDR OLED display",
                "screen_size": "6.1",
                "chip": "A16 Bionic",
                "camera": "48MP Main + 12MP Ultra Wide",
                "front_camera": "12MP TrueDepth front camera",
                "battery": "Video playback up to 20 hours",
                "connector": "USB-C",
                "network": "5G",
            }
        if version == 14 and is_pro:
            return {
                "screen": ("6.7-inch" if is_max else "6.1-inch") + " Super Retina XDR OLED display with ProMotion",
                "screen_size": "6.7" if is_max else "6.1",
                "chip": "A16 Bionic",
                "camera": "48MP Main + Ultra Wide + Telephoto",
                "front_camera": "12MP TrueDepth front camera",
                "battery": "Video playback up to 29 hours" if is_max else "Video playback up to 23 hours",
                "connector": "Lightning",
                "network": "5G",
            }
        if version == 13 and is_pro:
            return {
                "screen": ("6.7-inch" if is_max else "6.1-inch") + " Super Retina XDR OLED display with ProMotion",
                "screen_size": "6.7" if is_max else "6.1",
                "chip": "A15 Bionic",
                "camera": "12MP Pro camera system",
                "front_camera": "12MP TrueDepth front camera",
                "battery": "Video playback up to 28 hours" if is_max else "Video playback up to 22 hours",
                "connector": "Lightning",
                "network": "5G",
            }
        return {
            "screen": "6.1-inch Super Retina XDR OLED display",
            "screen_size": "6.1",
            "chip": "A15 Bionic",
            "camera": "12MP Dual camera system",
            "front_camera": "12MP TrueDepth front camera",
            "battery": "Video playback up to 19 hours",
            "connector": "Lightning",
            "network": "5G",
        }

    def image_urls_for(self, model, color_name):
        image_ids = self.image_ids_for(model, color_name)
        return [store_image(image_id) for image_id in image_ids]

    def image_ids_for(self, model, color_name):
        slug = self.color_slug_for_images(model, color_name)
        if model == "iPhone Air":
            return [
                f"iphone-air-finish-select-{slug}-202509",
                f"iphone-air-finish-select-{slug}-202509_AV2",
                f"iphone-air-finish-select-{slug}-202509_AV3",
            ]
        if model == "iPhone 17":
            return [
                f"iphone-17-finish-select-{slug}-202509_GEO_US",
                f"iphone-17-finish-select-{slug}-202509_AV2",
                f"iphone-17-finish-select-{slug}-202509_AV3",
            ]
        if model in {"iPhone 17 Pro", "iPhone 17 Pro Max"}:
            prefix = "iphone-17-pro-max" if model.endswith("Max") else "iphone-17-pro"
            return [
                f"{prefix}-finish-select-{slug}-202509",
                f"{prefix}-finish-select-{slug}-202509_AV2",
                f"{prefix}-finish-select-{slug}-202509_AV3",
            ]
        if model == "iPhone 16e":
            return [
                f"iphone-16e-finish-select-202502-{slug}",
                f"iphone-16e-finish-select-202502-{slug}_AV1",
                f"iphone-16e-finish-select-202502-{slug}_AV2",
            ]
        if model in {"iPhone 16", "iPhone 16 Plus"}:
            prefix = "iphone-16-plus" if model.endswith("Plus") else "iphone-16"
            return [
                f"{prefix}-{slug}-select-202409",
                f"{prefix}-{slug}-select-202409_AV2",
                f"{prefix}-{slug}-select-202409_AV3",
            ]
        if model in {"iPhone 16 Pro", "iPhone 16 Pro Max"}:
            size = "6-9inch" if model.endswith("Max") else "6-3inch"
            return [
                f"iphone-16-pro-finish-select-202409-{size}-{slug}",
                f"iphone-16-pro-finish-select-202409-{size}-{slug}_AV1",
                f"iphone-16-pro-finish-select-202409-{size}-{slug}_AV3",
            ]
        if model in {"iPhone 15 Pro", "iPhone 15 Pro Max"}:
            size = "6-7inch" if model.endswith("Max") else "6-1inch"
            return [
                f"iphone-15-pro-finish-select-202309-{size}-{slug}",
                f"iphone-15-pro-finish-select-202309-{size}-{slug}_AV1",
                f"iphone-15-pro-finish-select-202309-{size}-{slug}_AV2",
            ]
        if model == "iPhone 15":
            return [
                f"iphone-15-finish-select-202309-6-1inch-{slug}",
                f"iphone-15-finish-select-202309-6-1inch-{slug}_AV1",
                f"iphone-15-finish-select-202309-6-1inch-{slug}_AV2",
            ]
        if model in {"iPhone 14 Pro", "iPhone 14 Pro Max"}:
            size = "6-7inch" if model.endswith("Max") else "6-1inch"
            return [
                f"iphone-14-pro-finish-select-202209-{size}-{slug}",
                f"iphone-14-pro-finish-select-202209-{size}-{slug}_AV1",
                f"iphone-14-pro-finish-select-202209-{size}-{slug}_AV2",
            ]
        if model.startswith("iPhone 13"):
            base_slug = slug if slug in {"green", "pink", "blue", "midnight", "starlight"} else "blue"
            return [
                f"iphone-13-finish-select-202207-{base_slug}",
                f"iphone-13-finish-select-202207-{base_slug}_AV1",
                f"iphone-13-finish-select-202207-{base_slug}_AV2",
            ]
        return [
            "iphone-16-black-select-202409",
            "iphone-16-black-select-202409_AV2",
            "iphone-16-black-select-202409_AV3",
        ]

    def color_slug_for_images(self, model, color_name):
        mapping = {
            "Black": "black",
            "Jet Black": "black",
            "White": "white",
            "Pink": "pink",
            "Teal": "teal",
            "Ultramarine": "ultramarine",
            "Blue": "blue",
            "Deep Blue": "deepblue",
            "Sky Blue": "skyblue",
            "Mist Blue": "mistblue",
            "Green": "green",
            "Sage": "sage",
            "Deep Purple": "deeppurple",
            "Silver": "silver",
            "Cosmic Orange": "cosmicorange",
            "Desert Titanium": "deserttitanium",
            "Natural Titanium": "naturaltitanium",
            "Space Black": "blacktitanium",
        }
        slug = mapping.get(color_name, "black")
        if model in {"iPhone 15 Pro", "iPhone 15 Pro Max"} and color_name in {"Black", "Space Black"}:
            return "blacktitanium"
        if model in {"iPhone 15 Pro", "iPhone 15 Pro Max"} and color_name == "Blue Titanium":
            return "bluetitanium"
        if model in {"iPhone 15 Pro", "iPhone 15 Pro Max"} and color_name == "White Titanium":
            return "whitetitanium"
        if model in {"iPhone 16 Pro", "iPhone 16 Pro Max"} and color_name in {"Black", "Space Black"}:
            return "blacktitanium"
        if model in {"iPhone 16 Pro", "iPhone 16 Pro Max"} and color_name == "White Titanium":
            return "whitetitanium"
        if model in {"iPhone 14 Pro", "iPhone 14 Pro Max"} and color_name in {"Black", "Space Black"}:
            return "spaceblack"
        if model == "iPhone 17" and color_name == "Blue":
            return "mistblue"
        if model == "iPhone Air" and color_name == "Space Black":
            return "spaceblack"
        if model == "iPhone 16e" and slug not in {"black", "white"}:
            return "white"
        return slug

    def sku_from_key(self, key):
        cleaned = re.sub(r"[^A-Z0-9]+", "-", key).strip("-")
        digest = hashlib.sha1(key.encode("utf-8")).hexdigest()[:6].upper()
        return f"{cleaned[:82]}-{digest}"

    def upsert_product(self, item, category, brand, brand_category, color, memory):
        product = ProductVariant.objects.filter(sku=item["sku"]).select_related("product").first()
        product = product.product if product else Product.objects.filter(name=item["name"]).first()
        if product is None:
            product = Product(name=item["name"])
        product.category = category.name
        product.category_ref = category
        product.brand_name = brand.name
        product.brand_ref = brand
        product.brand_category = brand_category.name
        product.brand_category_ref = brand_category
        product.description = item["description"]
        product.save()
        product.colors.add(color)
        product.memories.add(memory)
        return product

    def upsert_variant(self, item, product, color, memory):
        specs = item["specs"]
        attrs = {
            "Тип": "Смартфон",
            "Производители": "Apple",
            "Модель": item["model"],
            "Память": item["storage"],
            "Цвет": color.name,
            "Состояние": item["condition"],
            "Экран": specs["screen"],
            "Диагональ экрана": specs["screen_size"],
            "Процессор": specs["chip"],
            "Камера": specs["camera"],
            "Фронтальная камера": specs["front_camera"],
            "Связь": specs["network"],
            "Разъем": specs["connector"],
            "Операционная система": "iOS",
            "Аккумулятор": specs["battery"],
        }
        if item["region"]:
            attrs["Регион"] = item["region"]
        variant, _ = ProductVariant.objects.update_or_create(
            sku=item["sku"],
            defaults={
                "product": product,
                "color": color,
                "memory": memory,
                "attributes": attrs,
                "is_active": True,
            },
        )
        return variant

    def upsert_stock(self, variant, shop, quantity, usd_price):
        Stock.objects.update_or_create(
            variant=variant,
            shop=shop,
            defaults={
                "quantity": int(quantity),
                "in_stock": quantity > 0,
                "wholesale_price": usd_price,
            },
        )

    def upsert_price(self, variant, channel, price):
        price_obj, _ = ChannelPrice.objects.update_or_create(
            variant=variant,
            shop=channel.shop,
            channel=channel,
            defaults={
                "price": price,
                "sync_status": ChannelPrice.SyncStatus.PENDING,
                "last_sync_error": "",
            },
        )
        return price_obj

    def ensure_images(self, variant, item):
        existing_images = list(variant.images.all())
        has_valid_images = len(existing_images) >= 3 and all(
            urlparse(image.image.url).path.lower().endswith((".jpg", ".jpeg", ".png", ".webp"))
            for image in existing_images
        )
        if has_valid_images:
            return
        variant.images.all().delete()
        for index, image_url in enumerate(item["image_urls"]):
            content = self.download_image(image_url, f"{item['sku'].lower()}_{index}.jpg")
            ProductImage.objects.create(
                variant=variant,
                image=content,
                color=variant.color,
                is_primary=index == 0,
                order=index,
            )

    def download_image(self, url, filename):
        last_error = None
        for _ in range(3):
            try:
                response = requests.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=45)
                response.raise_for_status()
                image = Image.open(BytesIO(response.content)).convert("RGBA")
                break
            except (requests.RequestException, OSError) as exc:
                last_error = exc
        else:
            raise CommandError(f"Фото {url} недоступно: {last_error}") from last_error

        canvas = Image.new("RGBA", image.size, "WHITE")
        canvas.alpha_composite(image)
        rgb_image = canvas.convert("RGB")
        output = BytesIO()
        rgb_image.save(output, format="JPEG", quality=92, optimize=True)
        return ContentFile(output.getvalue(), name=filename)
