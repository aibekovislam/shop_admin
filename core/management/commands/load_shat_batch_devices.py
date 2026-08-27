from collections import OrderedDict
from decimal import Decimal, ROUND_HALF_UP
from io import BytesIO
import hashlib
import json
import re
from pathlib import Path
from urllib.parse import urlparse

from django.core.files.base import ContentFile
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook
from PIL import Image
import requests

from core.marketplace.factory import get_marketplace_adapter
from core.models import (
    Brand,
    BrandCategory,
    Channel,
    ChannelPrice,
    Memory,
    Product,
    ProductCategory,
    ProductColor,
    ProductImage,
    ProductSize,
    ProductVariant,
    Stock,
)


COLORS = {
    "BLACK": ("Black", "#1D1D1F"),
    "MIDNIGHT": ("Midnight", "#1D1D1F"),
    "SPACE GREY": ("Space Gray", "#6E6E73"),
    "SPACE GRAY": ("Space Gray", "#6E6E73"),
    "S/BLACK": ("Space Black", "#1D1D1F"),
    "SPACE BLACK": ("Space Black", "#1D1D1F"),
    "SILVER": ("Silver", "#D7D7D7"),
    "WHITE": ("White", "#F5F5F5"),
    "STARLIGHT": ("Starlight", "#F0E4D3"),
    "SKYBLUE": ("Sky Blue", "#AFCFEA"),
    "SKY BLUE": ("Sky Blue", "#AFCFEA"),
    "GREEN": ("Green", "#8BC6A3"),
    "ЗЕЛЕН": ("Green", "#8BC6A3"),
    "ЗЕЛЁН": ("Green", "#8BC6A3"),
    "BLUE": ("Blue", "#4E83C2"),
    "PURPLE": ("Purple", "#7E5BA6"),
    "GOLD": ("Gold", "#D4AF37"),
    "ROSE GOLD": ("Rose Gold", "#E6B7A9"),
    "NATURAL": ("Natural Titanium", "#C7BEB2"),
    "SLATE": ("Slate", "#4B5256"),
    "GREY": ("Grey", "#777B80"),
    "GRAY": ("Grey", "#777B80"),
    "BROWN": ("Brown", "#8B5E3C"),
    "BEIGE": ("Beige", "#D8CAB4"),
    "БЕЖЕВ": ("Beige", "#D8CAB4"),
    "ЧЕРН": ("Black", "#1D1D1F"),
    "БЕЛ": ("White", "#F5F5F5"),
    "КРАСН": ("Red", "#B3261E"),
}


IMAGE_SETS = {
    "airtag": [
        "https://static.k-tuin.com/media/catalog/product/cache/1/image/0dc2d03fe217f8c83829496872af24a0/c/o/comprar-airtag-apple_1.jpg",
        "https://sopiguard.com/cdn/shop/products/airtagcleargl.jpg?v=1620108733",
        "https://www.isetos.cz/out/pictures/z2/acosapaita010_max_1_z2.jpg",
    ],
    "apple_watch": [
        "https://bizweb.dktcdn.net/100/116/615/products/a-nh-chu-p-ma-n-hi-nh-2024-09-14-lu-c-15-36-48.png?v=1747315314587",
        "https://static.tecnichenuove.it/01smartlife/2024/09/Apple-Watch-Series-10.jpg",
        "https://bizweb.dktcdn.net/thumb/1024x1024/100/116/615/products/a-nh-ma-n-hi-nh-2024-09-10-lu-c-05-09-01.png",
        "https://asset.conrad.com/media10/isa/160267/c1/-/de/003329413PI00/image.jpg?align=center&ex=400&ey=400&format=jpg&x=400&y=400",
        "https://a.scdn.gr/images/sku_images/105531/105531588/20241219122511_cd037878.jpeg",
    ],
    "apple_watch_ultra": [
        "https://asset.conrad.com/media10/isa/160267/c1/-/de/003329413PI00/image.jpg?align=center&ex=400&ey=400&format=jpg&x=400&y=400",
        "https://a.scdn.gr/images/sku_images/105531/105531588/20241219122511_cd037878.jpeg",
        "https://static.tecnichenuove.it/01smartlife/2024/09/Apple-Watch-Series-10.jpg",
    ],
    "macbook": [
        "https://myistore.co.zm/cdn/shop/files/MacBook_Air_15-inch_M4_Starlight_PDP_Image_Position_1__WWEN_1e23dfa5-42f5-4e75-a279-71ff245261a9.jpg?v=1744708769",
        "https://f.nooncdn.com/p/pnsku/N70154906V/45/_/1764242027/5f9d7966-2e15-4f54-86c0-13cafb9a9818.jpg?width=480",
        "https://www.istorm.gr/cdn/shop/files/IMG-16746936_61f125b6-c663-4e30-955a-5a451c4f9ad7.jpg?v=1741293903&width=823",
    ],
    "macbook_pro": [
        "https://m.media-amazon.com/images/I/61rknT42j%20L._UF1000%2C1000_QL80_.jpg",
        "https://akm-img-a-in.tosshub.com/indiatoday/images/story/202410/macbook-pro-2023-292233844-16x9.jpg?VersionId=aeIO9F5larPkDJUOvnBrC1xi7oEpmQ0g",
        "https://store.storeimages.cdn-apple.com/4982/as-images.apple.com/is/mbp14-spaceblack-select-202410?wid=1000&hei=1000&fmt=jpeg&qlt=90&.v=1728916305295",
    ],
    "imac": [
        "https://images.tcdn.com.br/img/img_prod/757873/apple_imac_24_m4_verde_16gb_ram_e_256gb_ssd_potencia_e_desempenho_avancado_5333_1_94d955d094c0be377fac3966d967ef0b.jpg",
        "https://media.power-cdn.net/images/h-00d330eaa1870d5d7828d4da6281c6ee/products/3741323/3741323_7_600x600_t_g.webp",
        "https://heise.cloudimg.io/bound/1200x1200/q85.png-lossy-85.webp-lossy-85.foil1/_www-heise-de_/imgs/18/4/7/0/3/8/0/5/e7e8fe5ef24940e0f2d4ecef3a33f4edcf6b1c7561fa1ce6ff477df9532220ce-cfab604131efa984.jpeg",
        "https://hnau.imgix.net/media/catalog/product/i/m/imac_m4_2_port_green_pdp_image_position_3__au.jpg",
    ],
    "mac_mini": [
        "https://melrosemac.com/cdn/shop/files/c23d4af36aba471b4d7a9faa92f749b5_e6570d22-8eb6-4eeb-9492-f0b1ca023761.jpg?v=1748988298",
        "https://avatars.mds.yandex.net/get-mpic/5286714/2a00000193ea91e8a1fe08de90e48aaf38c2/orig",
        "https://store.storeimages.cdn-apple.com/4982/as-images.apple.com/is/mac-mini-202410-gallery-1?wid=1000&hei=1000&fmt=jpeg&qlt=90&.v=1728343711930",
    ],
    "xiaomi_phone": [
        "https://cdn.kalvo.com/uploads/img/gallery/80975-xiaomi-redmi-note-15-3.jpg",
        "https://a.allegroimg.com/original/29c85f/ff3558724c47862ed703b3363c7b/Xiaomi-Redmi-A5-Smartfon-4GB-128GB-Midnight-Black",
        "https://storage.micromagma.ma/micromagma/634bcd2f-0e8d-419f-904f-3ed9801f5c17.jpg",
        "https://assets.mmsrg.com/isr/166325/c1/-/ASSET_MMS_167751568?align=center&cdx=536&cdy=402&cox=0&coy=0&ex=536&ey=402&format=jpg&quality=80&resizesource=&sp=yes&strip=yes&trim=&unsharp=1.5x1+0.7+0.02&x=536&y=402",
    ],
    "poco_phone": [
        "https://jinglestore.ru/d/4b4d160e9cc2d19d74c7e56c0554fbb2a61ed090943bfe4a1484b5255f2dcb90.jpg",
        "https://cdn.nieuwemobiel.nl/media/poco/poco-m8-pro-4x3-64259dd2.webp",
        "https://www.mi-il.co.il/images/site/products/bbe13b5e-76bb-4e14-ba3a-7a81e881b586.jpg",
        "https://ekb.hi-stores.ru/upload/iblock/ed5/wp1vf53e91q2nix3j42cdbtupr7qhaqi.jpg",
    ],
    "xiaomi_tablet": [
        "https://images.everyeye.it/img-cover/xiaomi-redmi-pad-2-pro-v2-55234.jpg",
        "https://m.media-amazon.com/images/I/41Fysjq7D8L._AC_SX569_.jpg",
        "https://www.purepc.pl/image/news/2025/09/23_redmi_pad_2_pro_seria_tabletow_z_duzym_ekranem_i_bateria_12_000_mah_dostepny_matowy_wyswietlacz_5g_i_dolby_atmos_1_b.png",
    ],
    "xiaomi_air_purifier": [
        "https://rokbucket.rokomari.io/ProductNew20190903/1104X1581/Xiaomi_MIJIA_Air_Purifier_5_with_Support-Xiaomi-16cb3-444338.png",
        "https://img.drz.lazcdn.com/g/kf/S14cc0aa2dc6744ec817a39a1ff332c92O.jpg_720x720q80.jpg",
        "https://static.lvengine.net/busntech/Imgs/produtos/product_35729/B6445.jpg",
        "https://mi-store.pl/hpeciai/a5999566bb5a9e91d9f9e7926b07d757/eng_pl_Air-Purifier-with-ionizer-Xiaomi-Mi-Smart-Air-Purifier-4-1773_3.webp",
    ],
    "xiaomi_band": [
        "https://cdn.media.amplience.net/i/xcite/664781-02?fmt=auto&img404=default&qlt=75&w=2048",
        "https://cdn.pacifiko.com/image/cache/catalog/p/OWNlMzU3Mj_3-1000x1000.png",
        "https://www.mi-il.co.il/images/site/products/bab1df06-b876-4345-945a-856fa9923500.jpg",
    ],
    "huawei_watch": [
        "https://kainos-img.dgn.lt/photos2_25_495544133/img.jpg",
        "https://cdn2.37left.lk/images/huawei-watch-gt-6-pro-46mm-cZhOSPMl0leP.webp",
        "https://imgcdn.myt.mu/refresh/devices/deriv_561_2_69089159afe87.jpg",
        "https://www.trikart.com/media/catalog/product/h/u/huawei_watch_gt_6_46mm_-_green-3.jpg?auto=webp&quality=90&width=2500",
    ],
    "yandex_station": [
        "https://www.ixbt.com/img/x780x600/r30/00/02/88/38/yast-1.jpg",
        "https://media.ixbt.site/fit-in/587x600/https%3A/www.ixbt.com/img/r30/00/02/88/38/yast-8.jpg",
        "https://www.ixbt.com/img/x780x600/r30/00/02/88/38/yast-3.jpg",
    ],
    "yandex_station_max": [
        "https://cdn.lemanapro.ru/lmru/image/upload/f_auto/q_auto/dpr_1.0/c_pad/w_1000/h_1000/v1707731783/lmcode/Clgjs-oKFkCOMnIRQIIJIQ/89345167_tmp.jpg",
        "https://neocomputer.md/image/cache/catalog/products/ui-products/293c9672-4d7f-11ea-b816-00155d1de702/1f0c2395-3d39-4560-a86e-3e4d5c5842e8-800x800.png",
        "https://www.ixbt.com/img/r30/00/02/68/02/ya-18.jpg",
    ],
    "yandex_station_mini": [
        "https://hi-stores.ru/upload/iblock/4e6/v1h4w8ep5es9lvw664dkckunt0vnjubo.webp",
        "https://ultra-smart.ru/center/iblock/1dd/35068h4wiowbycqa5cjoim8nf8xljxxp/424fdea25975188f23f690bfc15eb30f7ef710250adb7466bc1d9df19b57d3d8.png.webp",
        "https://cdn.evrika.com/storage/products/images/big/s1NXLLZzRdKx7XQrKu8mFNsolrAZETVBU1hx9hwd.jpeg.webp",
    ],
}


class Command(BaseCommand):
    help = "Загружает выбранный большой батч SHAT из Excel Turan в M-Market."

    def add_arguments(self, parser):
        parser.add_argument("--excel", default="Price Turan .xlsx", help="Путь к Excel-файлу с остатками Turan.")
        parser.add_argument("--channel-id", type=int, help="ID канала M-Market.")
        parser.add_argument("--adapter-key", default="mmarket", help="Ключ адаптера M-Market.")
        parser.add_argument("--markup", type=Decimal, default=Decimal("0.15"), help="Наценка, например 0.15.")
        parser.add_argument("--usd-rate", type=Decimal, default=Decimal("88"), help="Курс USD/KGS.")
        parser.add_argument("--dry-run", action="store_true", help="Создать товары и показать payload без отправки.")
        parser.add_argument("--skip-images", action="store_true", help="Не скачивать фото, если их уже минимум 3.")

    def handle(self, *args, **options):
        channel = self.get_channel(options)
        rows = self.read_selected_rows(options["excel"])
        payload_ids = []
        self._download_cache = {}

        with transaction.atomic():
            for key, stock_row in rows.items():
                item = self.build_item(key)
                brand, _ = Brand.objects.get_or_create(name=item["brand"])
                brand_category, _ = BrandCategory.objects.get_or_create(brand=brand, name=item["brand_category"])
                category, _ = ProductCategory.objects.get_or_create(name=item["category"])
                color_name, color_hex = item["color"]
                color, _ = ProductColor.objects.get_or_create(name=color_name, defaults={"hash_code": color_hex})
                if color.hash_code != color_hex:
                    color.hash_code = color_hex
                    color.save(update_fields=["hash_code"])
                memory = None
                if item.get("memory"):
                    memory, _ = Memory.objects.get_or_create(volume=item["memory"])
                size = None
                if item.get("size"):
                    size, _ = ProductSize.objects.get_or_create(name=item["size"])

                product = self.upsert_product(item, category, brand, brand_category, color, memory, size)
                variant = self.upsert_variant(item, product, color, memory, size)
                wholesale_price = self.calculate_wholesale_price(stock_row["usd"], options["usd_rate"])
                self.upsert_stock(variant, channel.shop, stock_row["quantity"], wholesale_price)
                price = self.calculate_price(stock_row["usd"], options["usd_rate"], options["markup"])
                price_obj = self.upsert_price(variant, channel, price)
                payload_ids.append(price_obj.id)
                if not options["skip_images"]:
                    self.ensure_images(variant, item)

        adapter = get_marketplace_adapter(channel)
        payload = adapter.build_payload(channel_price_ids=payload_ids)
        self.stdout.write(f"Подготовлено товаров для M-Market: {len(payload['products'])}")
        self.stdout.write(f"ChannelPrice IDs: {payload_ids}")
        self.stdout.write(json.dumps(payload, ensure_ascii=False, indent=2))

        if options["dry_run"]:
            self.stdout.write(self.style.WARNING("Dry-run: товары созданы/обновлены, но запрос в M-Market не отправлен."))
            return

        result = adapter.push_products(channel_price_ids=payload_ids)
        self.stdout.write(self.style.SUCCESS(f"Отправлено в {channel}: {result}"))

    def get_channel(self, options):
        if options["channel_id"]:
            channel = Channel.objects.filter(id=options["channel_id"]).first()
        else:
            channel = Channel.objects.filter(adapter_key=options["adapter_key"], is_active=True).first()
        if not channel:
            raise CommandError("Канал не найден. Передай --channel-id или проверь adapter_key.")
        if not channel.api_url or not channel.api_token or not channel.branch_id:
            raise CommandError(f"У канала {channel} не заполнены API URL/API token/branch_id.")
        return channel

    def read_selected_rows(self, path):
        workbook = load_workbook(self.resolve_excel_path(path), data_only=True)
        sheet = workbook.active
        rows = OrderedDict()
        for row in sheet.iter_rows(values_only=True):
            name = row[0]
            if not name:
                continue
            key = self.normalize_excel_name(name)
            if key.startswith("ИТОГО") or self.is_already_loaded_category(key) or self.is_used(key):
                continue
            if not self.family_for(key):
                continue
            quantity = self.decimal_value(row[1]) or Decimal("0")
            usd = self.decimal_value(row[2])
            if not usd or quantity <= 0:
                continue
            entry = rows.setdefault(key, {"quantity": Decimal("0"), "usd": usd})
            entry["quantity"] += quantity
            entry["usd"] = max(entry["usd"], usd)
        return rows

    def resolve_excel_path(self, path):
        original_path = Path(path)
        candidates = [original_path]
        if not original_path.is_absolute():
            candidates.extend([Path.cwd() / original_path, Path("/app") / original_path, Path("/app/imports") / original_path])
        for candidate in candidates:
            if candidate.exists():
                return candidate
        checked = ", ".join(str(candidate) for candidate in candidates)
        raise CommandError(f"Excel-файл не найден. Проверенные пути: {checked}")

    def normalize_excel_name(self, value):
        return " ".join(str(value).upper().replace("Б/У", "B/U").replace("Б.У", "B/U").replace("Б\\У", "B/U").split())

    def decimal_value(self, value):
        if value is None:
            return None
        if isinstance(value, Decimal):
            return value
        if isinstance(value, (int, float)):
            return Decimal(str(value))
        return Decimal(str(value).replace(" ", "").replace(",", "."))

    def is_already_loaded_category(self, key):
        if "AIRPODS" in key or "AIR PODS" in key:
            return True
        if "IPAD" in key or key.startswith("APPLE PENCIL") or key.startswith("MAGIC KEYBOARD"):
            return True
        if "GARMIN" in key or "GARMINE" in key:
            return True
        if "IPHONE" in key:
            return True
        return bool(re.match(r"^(11|12|13|14|15|16|17)\s+(PRO|PROMAX|PRO MAX|PLUS|AIR|E|\d)", key))

    def is_used(self, key):
        return "B/U" in key or "DAMAGED" in key

    def calculate_price(self, usd_price, usd_rate, markup):
        return (usd_price * usd_rate * (Decimal("1") + markup)).quantize(Decimal("1"), rounding=ROUND_HALF_UP)

    def calculate_wholesale_price(self, usd_price, usd_rate):
        return (usd_price * usd_rate).quantize(Decimal("1"), rounding=ROUND_HALF_UP)

    def build_item(self, key):
        family = self.family_for(key)
        brand = self.brand_for(key, family)
        category = self.category_for(key, family)
        brand_category = self.brand_category_for(key, family)
        memory = self.extract_memory(key)
        size = self.extract_size(key)
        color = self.extract_color(key)
        model = self.extract_model(key, brand)
        name = self.title_name(key, brand)
        features = self.features_for(key, family)
        return {
            "sku": self.sku_from_key(key, brand),
            "name": name,
            "brand": brand,
            "brand_category": brand_category,
            "category": category,
            "model": model,
            "memory": memory,
            "size": size,
            "color": color,
            "type": brand_category,
            "features": features,
            "description": self.description_for(name, brand, category, features, color[0], memory, size),
            "image_urls": self.image_urls_for(key, family),
        }

    def family_for(self, key):
        if any(marker in key for marker in ("AIRTAG", "ADAPTER", "CHARGER", "CABLE", "MAGSAFE", "POWER ADAPTER", "USB-C", "LIGHTNING")):
            return "airtag" if "AIRTAG" in key else "apple_accessory"
        if key.startswith("APPLE WATCH") or key.startswith("AW ") or re.match(r"^(SE|SERIES|ULTRA)\b", key):
            return "apple_watch"
        if "HUAWEI" in key or "HONOR" in key:
            return "huawei_watch"
        if any(marker in key for marker in ("MACBOOK", "MAC BOOK", "IMAC", "MAC MINI", "MAC STUDIO", "MAC PRO")):
            return "mac"
        if any(marker in key for marker in ("XIAOMI", "XIAMOMI", "REDMI", "POCO")):
            return "xiaomi"
        if key.startswith("ЯНДЕКС") or "YANDEX" in key:
            return "yandex"
        return ""

    def brand_for(self, key, family):
        if family in {"airtag", "apple_accessory", "apple_watch", "mac"}:
            return "Apple"
        if family == "huawei_watch":
            return "Huawei"
        if family == "yandex":
            return "Яндекс"
        if "POCO" in key:
            return "POCO"
        return "Xiaomi"

    def category_for(self, key, family):
        if family == "airtag":
            return "Поисковые метки"
        if family == "apple_accessory":
            return "Аксессуары"
        if family in {"apple_watch", "huawei_watch"} or "SMART BAND" in key:
            return "Смарт-часы"
        if family == "mac":
            return "Ноутбуки" if "MACBOOK" in key or "MAC BOOK" in key else "Компьютеры"
        if family == "yandex":
            return "Умные колонки"
        if "PAD" in key:
            return "Планшеты"
        if "AIR PURIFIER" in key:
            return "Очистители воздуха"
        return "Смартфоны"

    def brand_category_for(self, key, family):
        if family == "airtag":
            return "AirTag"
        if family == "apple_accessory":
            return "Accessories"
        if family == "apple_watch":
            return "Apple Watch Ultra" if "ULTRA" in key else "Apple Watch"
        if family == "mac":
            if "IMAC" in key:
                return "iMac"
            if "MAC MINI" in key:
                return "Mac mini"
            return "MacBook Pro" if "PRO" in key else "MacBook Air"
        if family == "huawei_watch":
            return "Huawei Watch"
        if family == "yandex":
            return "Яндекс Станция"
        if "POCO" in key:
            return "POCO"
        if "PAD" in key:
            return "Redmi Pad"
        if "SMART BAND" in key:
            return "Smart Band"
        return "Redmi"

    def extract_model(self, key, brand):
        cleaned = key
        for token in ("XIAMOMI", "XIAOMI", "APPLE", "HUAWEI", "HONOR"):
            cleaned = cleaned.replace(token, "")
        return f"{brand} {self.title_model(cleaned.strip(' ,-'))}".strip()

    def title_name(self, key, brand):
        words = self.title_model(key.replace("XIAMOMI", "XIAOMI"))
        return words[:250]

    def title_model(self, value):
        keep_upper = {"M1", "M2", "M3", "M4", "M5", "GPS", "USB-C", "ZIGBEE", "GB", "TB", "RAM", "SSD", "GT6", "PRO", "SE", "AW"}
        result = []
        for word in value.split():
            stripped = word.strip(",()")
            if stripped in keep_upper or re.search(r"\d", stripped):
                result.append(word)
            else:
                result.append(word.capitalize())
        return " ".join(result)

    def extract_memory(self, key):
        match = re.search(r"(\d+)\s*/\s*(\d+GB|\d+TB)", key)
        if match:
            return match.group(2)
        match = re.search(r"\b(\d+GB|\d+TB|1TB|2TB|512GB|256GB|128GB)\b", key)
        return match.group(1) if match else ""

    def extract_size(self, key):
        if match := re.search(r"(\d{2}(?:\.\d)?)\s*(?:INCH|\")", key):
            return f"{match.group(1)} inch"
        if match := re.search(r"(\d{2})\s*MM", key):
            return f"{match.group(1)}mm"
        if "S/M" in key or "S-M" in key:
            return "S/M"
        if "M/L" in key:
            return "M/L"
        if "MEDIUM" in key:
            return "Medium"
        return ""

    def extract_color(self, key):
        for marker in sorted(COLORS, key=len, reverse=True):
            if marker in key:
                return COLORS[marker]
        return COLORS["BLACK"]

    def features_for(self, key, family):
        if family == "airtag":
            return ["поиск вещей через Find My", "Bluetooth", "сменная батарейка", "компактный корпус"]
        if family == "apple_watch":
            return ["Retina дисплей", "мониторинг здоровья", "тренировки", "уведомления", "совместимость с iPhone"]
        if family == "mac":
            return ["Apple Silicon", "macOS", "Retina дисплей" if "MACBOOK" in key or "IMAC" in key else "компактный корпус", "Wi-Fi", "Bluetooth"]
        if family == "huawei_watch":
            return ["AMOLED дисплей", "мониторинг здоровья", "спортивные режимы", "длительная автономность", "Bluetooth"]
        if family == "yandex":
            return ["Алиса", "Wi-Fi", "Bluetooth", "умный дом", "голосовое управление"]
        if "PAD" in key:
            return ["большой дисплей", "Android/HyperOS", "Wi-Fi", "Bluetooth", "тонкий корпус"]
        if "AIR PURIFIER" in key:
            return ["очистка воздуха", "датчик качества воздуха", "умное управление", "фильтрация", "тихая работа"]
        if "SMART BAND" in key:
            return ["AMOLED дисплей", "мониторинг здоровья", "спортивные режимы", "уведомления", "длительная автономность"]
        return ["Android/HyperOS", "камера", "емкий аккумулятор", "Wi-Fi", "Bluetooth"]

    def description_for(self, name, brand, category, features, color, memory, size):
        parts = [
            f"{name} — оригинальный товар {brand}, категория: {category}.",
            f"Основные возможности: {', '.join(features)}.",
            f"Цвет: {color}.",
        ]
        if memory:
            parts.append(f"Память: {memory}.")
        if size:
            parts.append(f"Размер: {size}.")
        parts.append("Состояние: Новый. Подходит для ежедневного использования.")
        return " ".join(parts)[:1000]

    def image_urls_for(self, key, family):
        if family == "apple_watch":
            return IMAGE_SETS["apple_watch_ultra"] if "ULTRA" in key else IMAGE_SETS["apple_watch"]
        if family == "mac":
            if "IMAC" in key:
                return IMAGE_SETS["imac"]
            if "MAC MINI" in key:
                return IMAGE_SETS["mac_mini"]
            if "PRO" in key:
                return IMAGE_SETS["macbook_pro"]
            return IMAGE_SETS["macbook"]
        if family == "xiaomi":
            if "POCO" in key:
                return IMAGE_SETS["poco_phone"]
            if "PAD" in key:
                return IMAGE_SETS["xiaomi_tablet"]
            if "AIR PURIFIER" in key:
                return IMAGE_SETS["xiaomi_air_purifier"]
            if "SMART BAND" in key:
                return IMAGE_SETS["xiaomi_band"]
            return IMAGE_SETS["xiaomi_phone"]
        if family == "yandex":
            if "МАКС" in key or "MAX" in key:
                return IMAGE_SETS["yandex_station_max"]
            if "МИНИ" in key or "MINI" in key:
                return IMAGE_SETS["yandex_station_mini"]
            return IMAGE_SETS["yandex_station"]
        return IMAGE_SETS.get(family, IMAGE_SETS["airtag"])

    def sku_from_key(self, key, brand):
        base = re.sub(r"[^A-ZА-Я0-9]+", "-", key.replace("ЯНДЕКС", "YANDEX")).strip("-")
        prefix = re.sub(r"[^A-Z0-9]+", "", brand.upper())[:8] or "ITEM"
        digest = hashlib.sha1(key.encode("utf-8")).hexdigest()[:6].upper()
        return f"{prefix}-{base[:74]}-{digest}"

    def upsert_product(self, item, category, brand, brand_category, color, memory, size):
        product = ProductVariant.objects.filter(sku=item["sku"]).select_related("product").first()
        product = product.product if product else Product.objects.filter(name=item["name"]).first()
        if product is None:
            product = Product(name=item["name"])
        product.category = category.name
        product.category_ref = category
        product.brand_name = brand.name
        product.brand_ref = brand
        product.brand_category = brand_category.name
        product.brand_category_ref = brand_category
        product.description = item["description"]
        product.save()
        product.colors.add(color)
        if memory:
            product.memories.add(memory)
        if size:
            product.sizes.add(size)
        return product

    def upsert_variant(self, item, product, color, memory, size):
        attrs = {
            "Тип": item["type"],
            "Производители": item["brand"],
            "Модель": item["model"],
            "Цвет": color.name,
            "Состояние": "Новый",
            "Особенности": ", ".join(item["features"]),
        }
        if memory:
            attrs["Память"] = memory.volume
        if size:
            attrs["Размер"] = size.name
        variant, _ = ProductVariant.objects.update_or_create(
            sku=item["sku"],
            defaults={
                "product": product,
                "color": color,
                "memory": memory,
                "size": size,
                "attributes": attrs,
                "is_active": True,
            },
        )
        return variant

    def upsert_stock(self, variant, shop, quantity, wholesale_price):
        Stock.objects.update_or_create(
            variant=variant,
            shop=shop,
            defaults={"quantity": int(quantity), "in_stock": quantity > 0, "wholesale_price": wholesale_price},
        )

    def upsert_price(self, variant, channel, price):
        price_obj, _ = ChannelPrice.objects.update_or_create(
            variant=variant,
            shop=channel.shop,
            channel=channel,
            defaults={"price": price, "sync_status": ChannelPrice.SyncStatus.PENDING, "last_sync_error": ""},
        )
        return price_obj

    def ensure_images(self, variant, item):
        existing_images = list(variant.images.all())
        if len(existing_images) >= 3 and all(urlparse(image.image.url).path.lower().endswith((".jpg", ".jpeg", ".png", ".webp")) for image in existing_images):
            return
        variant.images.all().delete()
        downloaded = 0
        errors = []
        for image_url in item["image_urls"]:
            try:
                content = self.download_image(image_url, f"{item['sku'].lower()}_{downloaded}.jpg")
            except CommandError as exc:
                errors.append(str(exc))
                continue
            ProductImage.objects.create(
                variant=variant,
                image=content,
                color=variant.color,
                is_primary=downloaded == 0,
                order=downloaded,
            )
            downloaded += 1
            if downloaded == 3:
                return
        raise CommandError(f"Для {item['sku']} скачано только {downloaded} фото. Ошибки: {'; '.join(errors[:3])}")

    def download_image(self, url, filename):
        cache = getattr(self, "_download_cache", {})
        if url in cache and cache[url]:
            return ContentFile(cache[url], name=filename)
        if url in cache and cache[url] is None:
            raise CommandError(f"Фото {url} уже проверено и недоступно")
        cached = cache.get(url)
        if cached:
            return ContentFile(cached, name=filename)
        last_error = None
        for _ in range(2):
            try:
                response = requests.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=8)
                response.raise_for_status()
                image = Image.open(BytesIO(response.content)).convert("RGBA")
                break
            except (requests.RequestException, OSError) as exc:
                last_error = exc
        else:
            if hasattr(self, "_download_cache"):
                self._download_cache[url] = None
            raise CommandError(f"Фото {url} недоступно: {last_error}") from last_error
        image.thumbnail((1200, 1200), Image.LANCZOS)
        canvas = Image.new("RGBA", image.size, "WHITE")
        canvas.alpha_composite(image)
        output = BytesIO()
        canvas.convert("RGB").save(output, format="JPEG", quality=92, optimize=True)
        content = output.getvalue()
        if hasattr(self, "_download_cache"):
            self._download_cache[url] = content
        return ContentFile(content, name=filename)
