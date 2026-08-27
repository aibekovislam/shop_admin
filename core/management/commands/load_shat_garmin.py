from collections import OrderedDict
from decimal import Decimal, ROUND_HALF_UP
from io import BytesIO
import hashlib
import json
import re
from pathlib import Path
from urllib.parse import urlparse

from django.core.files.base import ContentFile
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook
from PIL import Image
import requests

from core.marketplace.factory import get_marketplace_adapter
from core.models import (
    Brand,
    BrandCategory,
    Channel,
    ChannelPrice,
    Product,
    ProductCategory,
    ProductColor,
    ProductImage,
    ProductSize,
    ProductVariant,
    Stock,
)


GARMIN_BASE = "https://res.garmin.com/en/products"


COLORS = {
    "BLACK": ("Black", "#1D1D1F"),
    "GRAPHITE": ("Graphite", "#3E4244"),
    "CARBON GRAY": ("Carbon Gray", "#4B4E51"),
    "SLATE": ("Slate", "#4D5357"),
    "SILVER": ("Silver", "#D7D7D7"),
    "TITANIUM": ("Titanium", "#B7B0A4"),
    "LUNAR GOLD": ("Lunar Gold", "#D6C3A1"),
    "SOFT GOLD": ("Soft Gold", "#D4BD9B"),
    "CREAM GOLD": ("Cream Gold", "#D7BC8E"),
    "LILAC": ("Lilac", "#C9B4D7"),
    "WHITE": ("White", "#F5F5F5"),
    "WHITESTONE": ("Whitestone", "#ECE9E2"),
    "IVORY": ("Ivory", "#F1E7D8"),
    "BONE": ("Bone", "#DED6C7"),
    "MOSS": ("Moss", "#6F7765"),
    "BLUE": ("Captain Blue", "#1E4D75"),
    "CAPTAIN BLUE": ("Captain Blue", "#1E4D75"),
    "YELLOW": ("Amp Yellow", "#D4CF2D"),
    "AMBER": ("Amber", "#C58A45"),
    "NEO TROPIC": ("Neo Tropic", "#1DBFA7"),
    "POWDER GRAY": ("Powder Gray", "#BFC2C3"),
    "LICHEN": ("Lichen Camo", "#879278"),
    "CAMO": ("Camo", "#6F7566"),
}


LINE_PRODUCTS = {
    "DASH CAM": ("010-02231-10", "g"),
    "EDGE": ("010-02424-20", "v"),
    "ENDURO 3": ("010-02751-01", "v"),
    "ENDURO": ("010-02754-00", "v"),
    "EPIX": ("010-02582-10", "v"),
    "FENIX 6": ("010-02158-10", "v"),
    "FENIX 7": ("010-02540-20", "v"),
    "FENIX 8 PRO 51": ("010-03380-00", "v"),
    "FENIX 8 PRO": ("010-03198-00", "v"),
    "FENIX 8 43": ("010-02903-10", "v"),
    "FENIX 8 47": ("010-02904-10", "v"),
    "FENIX 8 51": ("010-02905-10", "v"),
    "FENIX": ("010-02904-10", "v"),
    "FORERUNNER 570 42": ("010-02970-00", "v"),
    "FORERUNNER 570": ("010-02971-00", "v"),
    "FORERUNNER 970": ("010-02969-00", "v"),
    "FORERUNNER 965": ("010-02809-10", "v"),
    "FORERUNNER 935": ("010-01746-00", "v"),
    "FORERUNNER 745": ("010-02445-10", "v"),
    "FORERUNNER 70": ("010-04623-00", "v"),
    "FORERUNNER 55": ("010-02562-10", "v"),
    "HRM": ("010-13388-00", "v"),
    "INDEX SLEEP": ("010-03024-00", "g"),
    "INDEX S2": ("010-02294-13", "g"),
    "INSTINCT 2": ("010-02626-10", "v"),
    "INSTINCT SOLAR": ("010-02293-16", "v"),
    "LILY": ("010-02839-02", "v"),
    "MARQ ADVENTURER": ("010-02648-00", "g"),
    "MARQ ATHLETE": ("010-02648-40", "g"),
    "MARQ COMMANDER": ("010-02648-10", "g"),
    "MARQ GOLFER": ("010-02648-20", "g"),
    "MARQ": ("010-02648-00", "g"),
    "QUATIX": ("010-02906-20", "v"),
    "TACTIX": ("010-02704-00", "v"),
    "VARIA": ("010-01674-00", "v"),
    "VENU 2": ("010-02496-11", "v"),
    "VENU 3S": ("010-02785-01", "v"),
    "VENU 4 45": ("010-03014-03", "v"),
    "VENU 4": ("010-03013-03", "v"),
    "VENU X1": ("010-02980-00", "v"),
    "VIVOACTIVE 5": ("010-02862-10", "v"),
    "VIVOACTIVE 6": ("010-02985-00", "v"),
    "VIVOMOVE": ("010-02239-01", "v"),
    "VIVOSMART": ("010-01995-00", "v"),
    "WATCH": ("010-02904-10", "v"),
}


SPECIAL_IMAGE_URLS = {
    "DASH CAM": [
        f"{GARMIN_BASE}/010-02231-10/g/26837-1-5623bb50-0ca2-4c41-8904-028f9f7f46de.jpg",
        f"{GARMIN_BASE}/010-02231-10/g/26837-2-812381d2-2079-470e-94e9-48eb4eb0db90.jpg",
        f"{GARMIN_BASE}/010-02231-10/g/26837-3-3bc3ba54-84b5-432c-bfc7-3fa512a5cd43.jpg",
    ],
    "EDGE": [
        f"{GARMIN_BASE}/010-02424-20/v/cf-lg-7cea0ad7-8d1b-4c9a-b819-52137378285c.jpg",
        f"{GARMIN_BASE}/010-02424-20/v/lf-lg-5b554954-c8f2-4380-9472-1aee7fc18e57.jpg",
        f"{GARMIN_BASE}/010-02424-20/v/pd-01-lg-8cd3e5fa-c2ce-4af6-a9e9-326578464876.jpg",
    ],
    "VARIA": [
        f"{GARMIN_BASE}/010-01674-00/v/cf-lg.jpg",
        f"{GARMIN_BASE}/010-01674-00/v/pd-01-lg.jpg",
        f"{GARMIN_BASE}/010-01674-00/v/pd-02-lg.jpg",
    ],
    "INDEX SLEEP": [
        f"{GARMIN_BASE}/010-03024-00/g/cf-lg.jpg",
        f"{GARMIN_BASE}/010-03024-00/g/rf-lg.jpg",
        f"{GARMIN_BASE}/010-03024-00/g/lf-lg.jpg",
    ],
    "INDEX S2": [
        f"{GARMIN_BASE}/010-02294-02/v/cf-lg-67d96e3b-7d7e-4c49-a9af-e53ff76dcc98.jpg",
        f"{GARMIN_BASE}/010-02294-02/v/rf-lg-bb73a9a4-55d6-4fbe-a34e-944f0c0ee299.jpg",
        f"{GARMIN_BASE}/010-02294-02/v/lf-lg-50251887-97a3-4eee-a7e9-2491749557e9.jpg",
    ],
    "ENDURO": [
        f"{GARMIN_BASE}/010-02754-00/g/cf-lg.jpg",
        f"{GARMIN_BASE}/010-02754-00/g/rf-lg-101c98eb-ca36-4c5e-932b-dfa1796c03b0.jpg",
        f"{GARMIN_BASE}/010-02754-00/g/lf-lg-5921c2ab-0a74-4641-a4ed-4ed1e5bd5dc2.jpg",
    ],
    "EPIX": [
        f"{GARMIN_BASE}/010-02804-00/v/cf-lg.jpg",
        f"{GARMIN_BASE}/010-02804-00/v/rf-lg.jpg",
        f"{GARMIN_BASE}/010-02804-00/v/lf-lg.jpg",
    ],
    "FENIX 6": [
        f"{GARMIN_BASE}/010-02158-00/g/cf-lg-a92acf0a-4c66-4e9e-8154-d7e1859bbde3.jpg",
        f"{GARMIN_BASE}/010-02158-00/g/rf-lg-e6be824d-905d-4396-9b08-88e0e97417a3.jpg",
        f"{GARMIN_BASE}/010-02158-00/g/lf-lg-14060d79-db74-4d83-817f-8f8671459650.jpg",
    ],
    "FENIX 7": [
        f"{GARMIN_BASE}/010-02539-01/v/cf-lg-2759aec9-b3e9-47b5-96ed-1a066f0af687.jpg",
        f"{GARMIN_BASE}/010-02539-01/v/rf-lg-eeb3ff07-3a97-4b0c-8e4f-fd6277dc30f7.jpg",
        f"{GARMIN_BASE}/010-02539-01/v/lf-lg-d6b19f15-65f2-44f7-8c2a-5b1662a380df.jpg",
    ],
    "FORERUNNER 55": [
        f"{GARMIN_BASE}/010-02562-01/v/cf-lg-26e853f6-2122-4df7-b72e-3c63be17c82e.jpg",
        f"{GARMIN_BASE}/010-02562-01/v/rf-lg-65969f22-0765-4166-80f0-037a6c6ee0d1.jpg",
        f"{GARMIN_BASE}/010-02562-01/v/lf-lg-2b840b55-8b4c-405f-9582-0237a3c7c738.jpg",
    ],
    "FORERUNNER 70": [
        f"{GARMIN_BASE}/010-03920-00/v/cf-lg.jpg",
        f"{GARMIN_BASE}/010-03920-00/v/rf-lg.jpg",
        f"{GARMIN_BASE}/010-03920-00/v/lf-lg.jpg",
    ],
    "FORERUNNER 745": [
        f"{GARMIN_BASE}/010-02445-00/v/cf-lg-176dc461-6f42-455e-a71c-7d2e9ed90f59.jpg",
        f"{GARMIN_BASE}/010-02445-00/v/pd-01-lg-8ddccd09-60c5-4385-92b2-4370253f7361.jpg",
        f"{GARMIN_BASE}/010-02445-00/v/pd-02-lg-acb1d7eb-953b-49b2-9a7b-54ff67a42fa3.jpg",
    ],
    "INSTINCT 2": [
        f"{GARMIN_BASE}/010-02563-02/g/cf-lg-a24be7d3-1a3b-4b9f-a1a6-e1704e4eb7b9.jpg",
        f"{GARMIN_BASE}/010-02563-02/g/rf-lg-33637b05-dab6-4f56-8b49-eef296753b7a.jpg",
        f"{GARMIN_BASE}/010-02563-02/g/lf-lg-d7d46052-a173-4dde-bc26-7dc973fb0896.jpg",
    ],
    "INSTINCT SOLAR": [
        f"{GARMIN_BASE}/010-02293-10/v/cf-lg-b31fcc72-a9a9-466b-9231-8e57a732e677-1.jpg",
        f"{GARMIN_BASE}/010-02293-10/v/rf-lg-71310c1e-7c55-4789-8298-452aed4c11a9.jpg",
        f"{GARMIN_BASE}/010-02293-10/v/lf-lg-fd11d55f-e334-40b8-a8a2-c0922b43365f.jpg",
    ],
    "TACTIX": [
        f"{GARMIN_BASE}/010-02931-02/g/cf-lg.png",
        f"{GARMIN_BASE}/010-02931-02/g/rf-lg.png",
        f"{GARMIN_BASE}/010-02931-02/g/lf-lg.png",
    ],
    "VENU 2": [
        f"{GARMIN_BASE}/010-02496-00/v/cf-lg-28560279-f9cf-4376-b5c8-7adc593c501a.jpg",
        f"{GARMIN_BASE}/010-02496-00/v/rf-lg-85b118f6-1cd9-4106-9391-e90d6025d0f9.jpg",
        f"{GARMIN_BASE}/010-02496-00/v/lf-lg-a193e28f-f6b9-4d0d-9feb-965bc52f4db5.jpg",
    ],
    "VIVOMOVE": [
        f"{GARMIN_BASE}/010-02239-00/v/cf-lg-11e27c39-8eaf-4731-9975-1225a94f5231-1.jpg",
        f"{GARMIN_BASE}/010-02239-00/v/rf-lg-f1b2cb39-12c9-4a9b-aeab-52f8de5b4b64.jpg",
        f"{GARMIN_BASE}/010-02239-00/v/lf-lg-3ff8a7d7-d391-4635-a3c4-84acc43f8bea.jpg",
    ],
    "VIVOSMART": [
        f"{GARMIN_BASE}/010-01995-10/v/cf-lg-f182b3fe-a3b4-48c2-81e9-7ee9d9f5abb8.jpg",
        f"{GARMIN_BASE}/010-01995-10/v/rf-lg-c9410daa-7d73-4965-9b5c-b881444870c6.jpg",
        f"{GARMIN_BASE}/010-01995-10/v/lf-lg-08f96920-7a55-400a-9d10-8c4bba1f64c3.jpg",
    ],
}


class Command(BaseCommand):
    help = "Загружает Garmin из Excel Turan в канал SHAT / M-Market."

    def add_arguments(self, parser):
        parser.add_argument(
            "--excel",
            default="Price Turan .xlsx",
            help="Путь к Excel-файлу с остатками Turan.",
        )
        parser.add_argument("--channel-id", type=int, help="ID канала M-Market.")
        parser.add_argument("--adapter-key", default="mmarket", help="Ключ адаптера канала.")
        parser.add_argument("--markup", type=Decimal, default=Decimal("0.15"), help="Наценка, например 0.15.")
        parser.add_argument("--usd-rate", type=Decimal, default=Decimal("88"), help="Курс USD/KGS.")
        parser.add_argument("--dry-run", action="store_true", help="Создать товары и показать payload без отправки.")
        parser.add_argument("--skip-images", action="store_true", help="Не скачивать фото, если их уже минимум 3.")

    def handle(self, *args, **options):
        channel = self.get_channel(options)
        rows = self.read_garmin(options["excel"])
        payload_ids = []

        with transaction.atomic():
            brand, _ = Brand.objects.get_or_create(name="Garmin")

            for key, stock_row in rows.items():
                item = self.build_item(key)
                brand_category, _ = BrandCategory.objects.get_or_create(brand=brand, name=item["brand_category"])
                category, _ = ProductCategory.objects.get_or_create(name=item["category"])
                price = self.calculate_price(stock_row["usd"], options["usd_rate"], options["markup"])
                wholesale_price = self.calculate_wholesale_price(stock_row["usd"], options["usd_rate"])

                color_name, color_hex = item["color"]
                color, _ = ProductColor.objects.get_or_create(name=color_name, defaults={"hash_code": color_hex})
                if color.hash_code != color_hex:
                    color.hash_code = color_hex
                    color.save(update_fields=["hash_code"])

                size = None
                if item.get("size"):
                    size, _ = ProductSize.objects.get_or_create(name=item["size"])

                product = self.upsert_product(item, category, brand, brand_category, color, size)
                variant = self.upsert_variant(item, product, color, size)
                self.upsert_stock(variant, channel.shop, stock_row["quantity"], wholesale_price)
                price_obj = self.upsert_price(variant, channel, price)
                payload_ids.append(price_obj.id)
                if not options["skip_images"]:
                    self.ensure_images(variant, item)

        adapter = get_marketplace_adapter(channel)
        payload = adapter.build_payload(channel_price_ids=payload_ids)
        self.stdout.write(f"Подготовлено товаров для M-Market: {len(payload['products'])}")
        self.stdout.write(f"ChannelPrice IDs: {payload_ids}")
        self.stdout.write(json.dumps(payload, ensure_ascii=False, indent=2))

        if options["dry_run"]:
            self.stdout.write(self.style.WARNING("Dry-run: товары созданы/обновлены, но запрос в M-Market не отправлен."))
            return

        result = adapter.push_products(channel_price_ids=payload_ids)
        self.stdout.write(self.style.SUCCESS(f"Отправлено в {channel}: {result}"))

    def get_channel(self, options):
        if options["channel_id"]:
            channel = Channel.objects.filter(id=options["channel_id"]).first()
        else:
            channel = Channel.objects.filter(adapter_key=options["adapter_key"], is_active=True).first()
        if not channel:
            raise CommandError("Канал не найден. Передай --channel-id или проверь adapter_key.")
        if not channel.api_url or not channel.api_token or not channel.branch_id:
            raise CommandError(f"У канала {channel} не заполнены API URL/API token/branch_id.")
        return channel

    def read_garmin(self, path):
        workbook = load_workbook(self.resolve_excel_path(path), data_only=True)
        sheet = workbook.active
        rows = OrderedDict()
        for row in sheet.iter_rows(values_only=True):
            name = row[0]
            if not name:
                continue
            key = self.normalize_excel_name(name)
            if "GARMIN" not in key and "GARMINE" not in key:
                continue
            quantity = self.decimal_value(row[1]) or Decimal("0")
            usd = self.decimal_value(row[2])
            if not usd:
                continue
            entry = rows.setdefault(key.replace("GARMINE", "GARMIN"), {"quantity": Decimal("0"), "usd": usd})
            entry["quantity"] += quantity
            entry["usd"] = max(entry["usd"], usd)
        return rows

    def resolve_excel_path(self, path):
        original_path = Path(path)
        candidates = [original_path]
        if not original_path.is_absolute():
            candidates.extend(
                [
                    Path.cwd() / original_path,
                    Path("/app") / original_path,
                    Path("/app/imports") / original_path,
                    Path("/tmp") / original_path,
                ]
            )
        for candidate in candidates:
            if candidate.exists():
                return candidate
        checked = ", ".join(str(candidate) for candidate in candidates)
        raise CommandError(f"Excel-файл не найден. Проверенные пути: {checked}")

    def normalize_excel_name(self, value):
        return " ".join(str(value).upper().replace("Б/У", "B/U").replace("Б.У", "B/U").replace("Б\\У", "B/U").split())

    def decimal_value(self, value):
        if value is None:
            return None
        if isinstance(value, Decimal):
            return value
        if isinstance(value, (int, float)):
            return Decimal(str(value))
        return Decimal(str(value).replace(" ", "").replace(",", "."))

    def calculate_price(self, usd_price, usd_rate, markup):
        return (usd_price * usd_rate * (Decimal("1") + markup)).quantize(Decimal("1"), rounding=ROUND_HALF_UP)

    def calculate_wholesale_price(self, usd_price, usd_rate):
        return (usd_price * usd_rate).quantize(Decimal("1"), rounding=ROUND_HALF_UP)

    def build_item(self, key):
        model = self.extract_model(key)
        condition = "Б/У" if "B/U" in key or "DAMAGED" in key else "Новый"
        size = self.extract_size(key)
        color = self.extract_color(key)
        product_type = self.product_type(key)
        category = self.category_for(product_type)
        brand_category = self.brand_category_for(product_type, model)
        sku = self.sku_from_key(key)
        features = self.features_for(product_type, model)
        name_parts = [model]
        if size and size not in model:
            name_parts.append(size)
        if color[0].lower() not in model.lower():
            name_parts.append(color[0])
        if condition == "Б/У":
            name_parts.append("B/U")
        name = " ".join(name_parts)
        description = (
            f"{name} — оригинальное устройство Garmin, состояние: {condition}. "
            f"Основные возможности: {', '.join(features)}. "
            f"Цвет: {color[0]}. Подходит для спорта, здоровья, навигации и ежедневного использования."
        )
        return {
            "sku": sku,
            "name": name[:250],
            "model": model,
            "category": category,
            "brand_category": brand_category,
            "type": product_type,
            "condition": condition,
            "size": size,
            "color": color,
            "features": features,
            "description": description[:1000],
            "image_urls": self.image_urls_for(key),
        }

    def extract_model(self, key):
        cleaned = key.replace("GARMIN", "").replace("GARMINE", "").strip(" ,-")
        cleaned = re.sub(r"\bB/U\b|\bUPON REQUEST\b|\bDAMAGED\b", "", cleaned).strip(" ,-")
        cleaned = re.sub(r"\s+", " ", cleaned)
        part_index = cleaned.find(" 010-")
        if part_index != -1:
            cleaned = cleaned[:part_index].strip(" ,-")
        return f"Garmin {self.title_model(cleaned)}"

    def title_model(self, value):
        keep_upper = {"AMOLED", "DLC", "GPS", "HRM", "EMEA", "S2", "UT", "X1", "WI-FI", "M-XL", "XS-S"}
        words = []
        for word in value.split():
            stripped = word.strip(",")
            if stripped in keep_upper or re.search(r"\d", stripped):
                words.append(word)
            else:
                words.append(word.capitalize())
        return " ".join(words)

    def extract_size(self, key):
        match = re.search(r"(\d{2})\s*MM|(\d{2})MM", key)
        if match:
            return f"{match.group(1) or match.group(2)}mm"
        if "M-XL" in key:
            return "M-XL"
        if "XS-S" in key or "S-M" in key:
            return "S-M"
        if "L-XL" in key:
            return "L-XL"
        return ""

    def extract_color(self, key):
        for marker in (
            "CARBON GRAY",
            "CAPTAIN BLUE",
            "LUNAR GOLD",
            "SOFT GOLD",
            "CREAM GOLD",
            "POWDER GRAY",
            "NEO TROPIC",
            "JET BLACK",
            "WHITESTONE",
            "GRAPHITE",
            "TITANIUM",
            "SILVER",
            "SLATE",
            "BLACK",
            "WHITE",
            "IVORY",
            "LILAC",
            "GREEN",
            "YELLOW",
            "AMBER",
            "BLUE",
            "MOSS",
            "CAMO",
            "BONE",
            "LICHEN",
        ):
            if marker in key:
                return COLORS[marker]
        return COLORS["BLACK"]

    def product_type(self, key):
        if "DASH CAM" in key:
            return "Видеорегистратор"
        if "EDGE" in key:
            return "Велокомпьютер"
        if "HRM" in key:
            return "Пульсометр"
        if "INDEX S2" in key:
            return "Умные весы"
        if "INDEX SLEEP" in key:
            return "Трекер сна"
        if "VARIA" in key:
            return "Велофара"
        return "Смарт-часы"

    def category_for(self, product_type):
        return {
            "Видеорегистратор": "Автоэлектроника",
            "Велокомпьютер": "Велоаксессуары",
            "Пульсометр": "Спортивные аксессуары",
            "Умные весы": "Умные весы",
            "Трекер сна": "Фитнес-трекеры",
            "Велофара": "Велоаксессуары",
        }.get(product_type, "Смарт-часы")

    def brand_category_for(self, product_type, model):
        if product_type != "Смарт-часы":
            return product_type
        match = re.match(r"Garmin ([A-Za-z]+)", model)
        return match.group(1) if match else "Smartwatch"

    def features_for(self, product_type, model):
        if product_type == "Видеорегистратор":
            return ["видеозапись дороги", "компактный корпус", "GPS", "подключение к Garmin Drive"]
        if product_type == "Велокомпьютер":
            return ["GPS-навигация", "карты и маршруты", "тренировочные метрики", "совместимость с датчиками"]
        if product_type == "Пульсометр":
            return ["точное измерение пульса", "передача данных на часы Garmin", "спортивные тренировки", "удобный ремень"]
        if product_type == "Умные весы":
            return ["измерение веса", "оценка состава тела", "Wi-Fi синхронизация", "Garmin Connect"]
        if product_type == "Трекер сна":
            return ["мониторинг сна", "HRV Status", "пульс и дыхание ночью", "до 7 дней работы"]
        if product_type == "Велофара":
            return ["яркий передний свет", "совместимость с велокомпьютерами Garmin", "несколько режимов", "зарядка USB"]
        if "Fenix" in model or "Epix" in model or "Tactix" in model or "Enduro" in model or "Quatix" in model:
            return ["AMOLED/GPS функции", "мультиспортивные режимы", "навигация и карты", "пульсометр", "Garmin Connect"]
        if "Forerunner" in model:
            return ["GPS для бега", "тренировочные метрики", "пульсометр", "AMOLED/спортивный дисплей", "Garmin Connect"]
        if "Venu" in model or "Vivoactive" in model:
            return ["здоровье и фитнес", "AMOLED дисплей", "пульсометр", "сон и восстановление", "Garmin Connect"]
        if "Lily" in model:
            return ["компактные женские смарт-часы", "мониторинг здоровья", "фитнес-функции", "Garmin Connect"]
        if "MARQ" in model or "Marq" in model:
            return ["премиальный корпус", "AMOLED дисплей", "спортивные профили", "навигация", "Garmin Connect"]
        return ["GPS", "спортивные режимы", "мониторинг здоровья", "смарт-уведомления", "Garmin Connect"]

    def image_urls_for(self, key):
        line_key = self.line_key_for(key)
        if line_key in SPECIAL_IMAGE_URLS:
            return SPECIAL_IMAGE_URLS[line_key]
        pid, folder = LINE_PRODUCTS[line_key]
        return [
            f"{GARMIN_BASE}/{pid}/{folder}/cf-lg.jpg",
            f"{GARMIN_BASE}/{pid}/{folder}/rf-lg.jpg",
            f"{GARMIN_BASE}/{pid}/{folder}/lf-lg.jpg",
            f"{GARMIN_BASE}/{pid}/{folder}/pd-01-lg.jpg",
            f"{GARMIN_BASE}/{pid}/{folder}/pd-02-lg.jpg",
        ]

    def line_key_for(self, key):
        normalized = key.replace("GARMINE", "GARMIN")
        checks = [
            ("DASH CAM", "DASH CAM"),
            ("EDGE", "EDGE"),
            ("ENDURO 3", "ENDURO 3"),
            ("ENDURO", "ENDURO"),
            ("EPIX", "EPIX"),
            ("FENIX 6", "FENIX 6"),
            ("FENIX 7", "FENIX 7"),
            ("FENIX 8 PRO 51", "FENIX 8 PRO 51"),
            ("FENIX PRO 8 51", "FENIX 8 PRO 51"),
            ("FENIX 8 PRO", "FENIX 8 PRO"),
            ("FENIX PRO 8", "FENIX 8 PRO"),
            ("FENIX 8 43", "FENIX 8 43"),
            ("FENIX 8 47", "FENIX 8 47"),
            ("FENIX 8 51", "FENIX 8 51"),
            ("FENIX", "FENIX"),
            ("FORERUNNER 570 - 42", "FORERUNNER 570 42"),
            ("FORERUNNER 570-47", "FORERUNNER 570"),
            ("FORERUNNER 570", "FORERUNNER 570"),
            ("FORERUNNER 970", "FORERUNNER 970"),
            ("FORERUNNER 965", "FORERUNNER 965"),
            ("FORERUNNER 935", "FORERUNNER 935"),
            ("FORERUNNER 745", "FORERUNNER 745"),
            ("FORERUNNER 70", "FORERUNNER 70"),
            ("FORERUNNER 55", "FORERUNNER 55"),
            ("HRM", "HRM"),
            ("INDEX SLEEP", "INDEX SLEEP"),
            ("INDEX S2", "INDEX S2"),
            ("INSTINCT 2", "INSTINCT 2"),
            ("INSTINCT SOLAR", "INSTINCT SOLAR"),
            ("LILY", "LILY"),
            ("MARQ ADVENTURER", "MARQ ADVENTURER"),
            ("MARQ ATHLETE", "MARQ ATHLETE"),
            ("MARQ COMMANDER", "MARQ COMMANDER"),
            ("MARQ GOLFER", "MARQ GOLFER"),
            ("MARQ", "MARQ"),
            ("QUATIX", "QUATIX"),
            ("TACTIX", "TACTIX"),
            ("VARIA", "VARIA"),
            ("VENU X1", "VENU X1"),
            ("VENU 2", "VENU 2"),
            ("VENU 3S", "VENU 3S"),
            ("VENU 4 45", "VENU 4 45"),
            ("VENU 4", "VENU 4"),
            ("VIVOACTIVE 5", "VIVOACTIVE 5"),
            ("VIVOACTIVE 6", "VIVOACTIVE 6"),
            ("VIVOMOVE", "VIVOMOVE"),
            ("VIVOSMART", "VIVOSMART"),
            ("WATCH", "WATCH"),
        ]
        for marker, line_key in checks:
            if marker in normalized:
                return line_key
        return "WATCH"

    def sku_from_key(self, key):
        part_number = re.search(r"\b010-\d{5}-\d{2}\b", key)
        if part_number:
            base = f"GARMIN-{part_number.group(0)}"
        else:
            base = re.sub(r"[^A-Z0-9]+", "-", key.replace("GARMINE", "GARMIN")).strip("-")
        digest = hashlib.sha1(key.encode("utf-8")).hexdigest()[:6].upper()
        return f"{base[:82]}-{digest}"

    def upsert_product(self, item, category, brand, brand_category, color, size):
        product = ProductVariant.objects.filter(sku=item["sku"]).select_related("product").first()
        product = product.product if product else Product.objects.filter(name=item["name"]).first()
        if product is None:
            product = Product(name=item["name"])
        product.category = category.name
        product.category_ref = category
        product.brand_name = brand.name
        product.brand_ref = brand
        product.brand_category = brand_category.name
        product.brand_category_ref = brand_category
        product.description = item["description"]
        product.save()
        product.colors.add(color)
        if size:
            product.sizes.add(size)
        return product

    def upsert_variant(self, item, product, color, size):
        attrs = {
            "Тип": item["type"],
            "Производители": "Garmin",
            "Модель": item["model"],
            "Цвет": color.name,
            "Состояние": item["condition"],
            "Особенности": ", ".join(item["features"]),
        }
        if size:
            attrs["Размер"] = size.name
        variant, _ = ProductVariant.objects.update_or_create(
            sku=item["sku"],
            defaults={
                "product": product,
                "color": color,
                "size": size,
                "attributes": attrs,
                "is_active": True,
            },
        )
        return variant

    def upsert_stock(self, variant, shop, quantity, usd_price):
        Stock.objects.update_or_create(
            variant=variant,
            shop=shop,
            defaults={
                "quantity": int(quantity),
                "in_stock": quantity > 0,
                "wholesale_price": usd_price,
            },
        )

    def upsert_price(self, variant, channel, price):
        price_obj, _ = ChannelPrice.objects.update_or_create(
            variant=variant,
            shop=channel.shop,
            channel=channel,
            defaults={
                "price": price,
                "sync_status": ChannelPrice.SyncStatus.PENDING,
                "last_sync_error": "",
            },
        )
        return price_obj

    def ensure_images(self, variant, item):
        existing_images = list(variant.images.all())
        has_valid_images = len(existing_images) >= 3 and all(
            urlparse(image.image.url).path.lower().endswith((".jpg", ".jpeg", ".png", ".webp"))
            for image in existing_images
        )
        if has_valid_images:
            return
        variant.images.all().delete()
        downloaded = 0
        errors = []
        for image_url in item["image_urls"]:
            try:
                content = self.download_image(image_url, f"{item['sku'].lower()}_{downloaded}.jpg")
            except CommandError as exc:
                errors.append(str(exc))
                continue
            ProductImage.objects.create(
                variant=variant,
                image=content,
                color=variant.color,
                is_primary=downloaded == 0,
                order=downloaded,
            )
            downloaded += 1
            if downloaded == 3:
                return
        raise CommandError(f"Для {item['sku']} скачано только {downloaded} фото. Ошибки: {'; '.join(errors[:3])}")

    def download_image(self, url, filename):
        last_error = None
        for _ in range(3):
            try:
                response = requests.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=45)
                response.raise_for_status()
                image = Image.open(BytesIO(response.content)).convert("RGBA")
                break
            except (requests.RequestException, OSError) as exc:
                last_error = exc
        else:
            raise CommandError(f"Фото {url} недоступно: {last_error}") from last_error

        canvas = Image.new("RGBA", image.size, "WHITE")
        canvas.alpha_composite(image)
        rgb_image = canvas.convert("RGB")
        output = BytesIO()
        rgb_image.save(output, format="JPEG", quality=92, optimize=True)
        return ContentFile(output.getvalue(), name=filename)
