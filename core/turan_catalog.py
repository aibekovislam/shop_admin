"""Read WEBSITE TURAN stock and match wholesale prices without fuzzy guesses."""

from collections import defaultdict
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
import re
import unicodedata
from pathlib import Path
import json

from openpyxl import load_workbook


USD_RATE = Decimal("88")
WHOLESALE_MARKUP = Decimal("1.15")
WEB_MARKUP = Decimal("1.10")
USED = re.compile(r"(?<!\w)(?:[БB]\s*[/\\.\-]?\s*[УU]|БУ)(?!\w)|DAMAGED", re.I)
NAME_ALIASES = {
    "AIRPODS4ANC": "AIRPODS4ACTIVENOISECANCELLATIONMXP93",
}


def number(value):
    if value is None or isinstance(value, bool):
        return None
    try:
        result = Decimal(str(value).replace("\xa0", "").replace(" ", "").replace(",", "."))
    except InvalidOperation:
        return None
    return result if result.is_finite() else None


def identity(name):
    text = unicodedata.normalize("NFKC", str(name)).upper()
    text = re.sub(r"\bAIR\s+PODS\b", "AIRPODS", text)
    key = re.sub(r"[^\w]", "", text)
    return NAME_ALIASES.get(key, key)


def selling_price(amount, currency, source):
    amount = number(amount)
    if amount is None or amount <= 0:
        raise ValueError("Price must be positive")
    if currency not in {"USD", "KGS"}:
        raise ValueError("Price currency must be USD or KGS")
    if source not in {"wholesale", "web"}:
        raise ValueError("Unknown price source")
    if source == "wholesale" and currency != "USD":
        raise ValueError("Wholesale workbook prices must be USD")
    base = amount * USD_RATE if currency == "USD" else amount
    return (base * (WHOLESALE_MARKUP if source == "wholesale" else WEB_MARKUP)).quantize(
        Decimal("1"), rounding=ROUND_HALF_UP
    )


def stock_rows(path, sheet="Остатки на 04.09.2026"):
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet]
        rows = []
        active = False
        declared_total = None
        for index, row in enumerate(worksheet.values, 1):
            name = str(row[0] or "").strip() if row else ""
            quantity = number(row[1]) if len(row) > 1 else None
            if name.upper() == "WEBSITE TURAN":
                if active or rows:
                    raise ValueError("More than one WEBSITE TURAN section")
                active = True
                declared_total = quantity
                continue
            if not active:
                continue
            if name.casefold() == "итого":
                total = sum(item["quantity"] for item in rows)
                if quantity != total or declared_total not in (None, total):
                    raise ValueError("WEBSITE TURAN stock total does not reconcile")
                return rows
            if not name:
                continue
            if quantity is None or quantity < 0 or quantity != quantity.to_integral_value():
                raise ValueError(f"Invalid stock quantity at row {index}: {name}")
            rows.append({"row": index, "name": name, "quantity": int(quantity)})
        raise ValueError("WEBSITE TURAN section with closing total not found")
    finally:
        workbook.close()


def audit_catalog(stock_path, price_path, sheet="Остатки на 04.09.2026"):
    prices = defaultdict(list)
    workbook = load_workbook(price_path, read_only=True, data_only=True)
    try:
        for worksheet in workbook:
            website_section = False
            for index, row in enumerate(worksheet.values, 1):
                if row and str(row[0] or "").strip().upper() == "WEBSITE TURAN":
                    website_section = True
                    continue
                if row and str(row[0] or "").strip().casefold() == "итого":
                    website_section = False
                    continue
                if len(row) < 3 or not row[0] or USED.search(str(row[0])):
                    continue
                usd = number(row[2])
                if usd is not None and usd > 0:
                    prices[identity(row[0])].append(
                        {"sheet": worksheet.title, "row": index, "name": str(row[0]), "usd": str(usd),
                         "website_turan": website_section}
                    )
    finally:
        workbook.close()
    items = []
    excluded = []
    for row in stock_rows(stock_path, sheet):
        if USED.search(row["name"]) or row["quantity"] == 0:
            excluded.append({**row, "reason": "used_or_damaged" if USED.search(row["name"]) else "zero_stock"})
            continue
        matches = prices.get(identity(row["name"]), [])
        preferred = [match for match in matches if match["website_turan"]] or matches
        amounts = {Decimal(match["usd"]) for match in preferred}
        item = {**row, "wholesale_matches": matches, "selected_wholesale_matches": preferred}
        if amounts and max(amounts) - min(amounts) <= Decimal("0.01"):
            selected_usd = max(amounts)
            item.update(price_status="matched", selected_usd=str(selected_usd),
                        rounding_difference=len(amounts) > 1,
                        price_kgs=str(selling_price(selected_usd, "USD", "wholesale")))
        else:
            item.update(price_status="conflicting_wholesale" if amounts else "needs_research", price_kgs=None)
        items.append(item)
    return {
        "stock_file": str(stock_path), "sheet": sheet, "price_file": str(price_path),
        "usd_rate": str(USD_RATE), "wholesale_markup_percent": 15, "web_markup_percent": 10,
        "summary": {
            "products": len(items), "quantity": sum(item["quantity"] for item in items),
            "matched": sum(item["price_status"] == "matched" for item in items),
            "needs_research": sum(item["price_status"] == "needs_research" for item in items),
            "conflicting_wholesale": sum(item["price_status"] == "conflicting_wholesale" for item in items),
            "excluded": len(excluded),
        },
        "items": items, "excluded": excluded,
    }


def fill_web_prices(audit, path=None):
    path = Path(path) if path else Path(__file__).parent / "catalog_data" / "turan_web_prices.json"
    data = json.loads(path.read_text(encoding="utf-8"))
    offers = {}
    for offer in data["offers"]:
        if not offer["source"].startswith("https://"):
            raise ValueError("Для цены из интернета нужна ссылка на источник")
        for name in offer["names"]:
            key = identity(name)
            if key in offers:
                raise ValueError(f"Несколько интернет-цен для {name}")
            offers[key] = offer
    for item in audit["items"]:
        if item["price_status"] != "needs_research":
            continue
        offer = offers.get(identity(item["name"]))
        if offer:
            item.update(price_status="web", price_kgs=str(selling_price(offer["amount"], data["currency"], "web")),
                        web_price={"amount": offer["amount"], "currency": data["currency"],
                                   "source": offer["source"], "checked_at": data["checked_at"]})
    audit["summary"]["web"] = sum(item["price_status"] == "web" for item in audit["items"])
    audit["summary"]["needs_research"] = sum(item["price_status"] == "needs_research" for item in audit["items"])
    return audit
